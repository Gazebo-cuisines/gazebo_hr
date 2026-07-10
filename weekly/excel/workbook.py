from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Border, Side

from ..payroll_service import PayrollResult

_TABLE_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Appended category breakdown / grand total (matches monthly export layout).
_CATEGORY_COL = 2
_CATEGORY_NUM_START = 4
_CATEGORY_BAND_COLS = ["BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"]
_HOUR_BAND_COLS = ["BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"]
_EXCEL_NUM_FORMAT = "0.00"
_OVERALL_CATEGORY_ORDER = ("PROD", "PACK", "WRHS", "CLNR", "TECH", "OFFICE")

_ALL_DATA_SECTION_TITLES = (
    "Category breakdown (detailed)",
    "Category breakdown (grouped)",
    "EMP / Agency totals",
    "Category breakdown (overall)",
)


def _sum_hour_bands(rows: list[dict[str, Any]]) -> dict[str, float]:
    return {k: sum(float(r.get(k, 0.0) or 0.0) for r in rows) for k in _HOUR_BAND_COLS}


def build_emp_agency_total_df(result: PayrollResult) -> pd.DataFrame:
    """EMP (non-agency / Gazebo), AGENCY, and TOTAL sums for the five hour bands (section C)."""
    emp = _sum_hour_bands(result.gazebo_rows)
    agency = _sum_hour_bands(result.agency_rows)
    total = _sum_hour_bands(result.rows)
    return pd.DataFrame(
        [
            {"Category": "EMP", **emp},
            {"Category": "AGENCY", **agency},
            {"Category": "TOTAL", **total},
        ]
    )


def build_category_summary_hr_df(analysis_df: pd.DataFrame) -> pd.DataFrame:
    """Same data as Analysis with HR-friendly column names and a grand total row (section B)."""
    hr_cols = {
        "BasicHours": "Basic Hour",
        "MonFriOvertime": "Mon Fri O",
        "SatSunOvertime": "Sat Sun O",
        "AnnualHoliday": "Annual Ho",
        "TotalPaidHours": "Total Paid Hours",
    }
    if analysis_df.empty:
        return pd.DataFrame(columns=["Category", *list(hr_cols.values())])
    out = analysis_df.rename(columns=hr_cols)
    total_row: dict[str, Any] = {"Category": "Grand total"}
    for src, dst in hr_cols.items():
        total_row[dst] = float(analysis_df[src].sum())
    return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)


def build_hours_over_60_df(all_df: pd.DataFrame) -> pd.DataFrame:
    """Employees with TotalPaidHours strictly greater than 60."""
    if all_df.empty:
        return pd.DataFrame(columns=["Name", "Category", "SageNo", "TotalPaidHours", "BasicHours", "Overtime", "ContractedHours"])
    work = all_df.copy()
    work["_tp"] = pd.to_numeric(work["TotalPaidHours"], errors="coerce").fillna(0.0)
    filtered = work[work["_tp"] > 60.0].drop(columns=["_tp"], errors="ignore")
    cols = ["Name", "Category", "SageNo", "TotalPaidHours", "BasicHours", "Overtime", "ContractedHours"]
    have = [c for c in cols if c in filtered.columns]
    return filtered[have] if have else filtered


def _build_analysis_dataframe(all_df: pd.DataFrame) -> pd.DataFrame:
    if all_df.empty:
        return pd.DataFrame(columns=["Category", "BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"])
    return (
        all_df.groupby("Category", dropna=False)[["BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"]]
        .sum()
        .reset_index()
    )


def _overall_category_key(category: str) -> str:
    c = str(category).strip().upper()
    if "PKNG" in c or "DPCH" in c:
        return "PACK"
    if "PROD" in c:
        return "PROD"
    if "WRHS" in c:
        return "WRHS"
    if "CLNR" in c:
        return "CLNR"
    if "TECH" in c:
        return "TECH"
    if "OFCE" in c:
        return "OFFICE"
    return "OTHER"


def _build_grouped_analysis_dataframe(analysis_df: pd.DataFrame) -> pd.DataFrame:
    # Lazy: monthly_service imports from payroll_service facade (circular at module load).
    from ..monthly_service import _grouped_key

    if analysis_df.empty:
        return pd.DataFrame(columns=["Category", *_CATEGORY_BAND_COLS])
    grouped: dict[str, dict[str, float]] = {}
    for _, row in analysis_df.iterrows():
        k = _grouped_key(str(row["Category"]))
        if k not in grouped:
            grouped[k] = {c: 0.0 for c in _CATEGORY_BAND_COLS}
        for c in _CATEGORY_BAND_COLS:
            grouped[k][c] += float(row[c] or 0.0)
    return pd.DataFrame([{"Category": k, **grouped[k]} for k in sorted(grouped)])


def _build_overall_analysis_dataframe(analysis_df: pd.DataFrame) -> pd.DataFrame:
    if analysis_df.empty:
        return pd.DataFrame(columns=["Category", *_CATEGORY_BAND_COLS])
    buckets = {k: {c: 0.0 for c in _CATEGORY_BAND_COLS} for k in _OVERALL_CATEGORY_ORDER}
    for _, row in analysis_df.iterrows():
        b = _overall_category_key(str(row["Category"]))
        if b not in buckets:
            buckets[b] = {c: 0.0 for c in _CATEGORY_BAND_COLS}
        for c in _CATEGORY_BAND_COLS:
            buckets[b][c] += float(row[c] or 0.0)
    return pd.DataFrame([{"Category": k, **buckets[k]} for k in _OVERALL_CATEGORY_ORDER])


def build_overall_category_totals(rows: list[dict[str, Any]]) -> pd.DataFrame:
    """Roll per-category band sums into PROD/PACK/WRHS/CLNR/TECH/OFFICE (weekly + monthly)."""
    if not rows:
        return pd.DataFrame(columns=["Category", *_CATEGORY_BAND_COLS])
    return _build_overall_analysis_dataframe(pd.DataFrame(rows))


def _append_section_title(ws, start_row: int, title: str) -> int:
    """Write a section label in col B. Returns the next row for the table header."""
    ws.cell(start_row, _CATEGORY_COL, title)
    return start_row + 1


def _apply_table_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Apply thin borders to every cell in the given range (formatting only)."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = _TABLE_BORDER


def _band_totals_from_df(df: pd.DataFrame) -> dict[str, float]:
    return {c: float(df[c].sum()) for c in _CATEGORY_BAND_COLS}


def _band_totals_from_emp_agency_total(emp_agency_df: pd.DataFrame) -> dict[str, float]:
    total_row = emp_agency_df.loc[emp_agency_df["Category"] == "TOTAL"].iloc[0]
    return {c: float(total_row[c]) for c in _CATEGORY_BAND_COLS}


def _append_band_values_row(ws, start_row: int, label: str, bands: dict[str, float]) -> int:
    """Write a labeled row of hour-band values (cols B and D–H). Returns one past the last row."""
    ws.cell(start_row, _CATEGORY_COL, label)
    for i, col in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(start_row, _CATEGORY_NUM_START + i, float(bands[col]))
    return start_row + 1


def _append_grand_total_row_openpyxl(
    ws,
    totals_df: pd.DataFrame,
    start_row: int,
    *,
    label: str = "Grand total",
) -> int:
    """Write total label and column sums. Returns one past the last row written."""
    if totals_df.empty:
        return start_row
    r = start_row
    ws.cell(r, _CATEGORY_COL, label)
    for i, col in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(r, _CATEGORY_NUM_START + i, float(totals_df[col].sum()))
    return r + 1


def _append_hour_totals_block(
    ws,
    totals_df: pd.DataFrame,
    start_row: int,
    *,
    header_label: str = "Category",
    grand_label: str | None = "Grand total",
) -> int:
    """Write category/hour band table (cols B and D–H). Returns one past the last row."""
    if totals_df.empty:
        return start_row
    r = start_row
    ws.cell(r, _CATEGORY_COL, header_label)
    for i, h in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(r, _CATEGORY_NUM_START + i, h)
    r += 1
    for _, row in totals_df.iterrows():
        cat = row["Category"]
        ws.cell(r, _CATEGORY_COL, "" if pd.isna(cat) else str(cat))
        for i, h in enumerate(_CATEGORY_BAND_COLS):
            v = row[h]
            ws.cell(r, _CATEGORY_NUM_START + i, 0.0 if pd.isna(v) else float(v))
        r += 1
    if grand_label:
        r += 1
        return _append_grand_total_row_openpyxl(ws, totals_df, r, label=grand_label)
    return r


def _append_category_breakdown_block(ws, analysis_df: pd.DataFrame, start_row: int) -> int:
    """Granular category breakdown with grand total (tier 1)."""
    return _append_hour_totals_block(ws, analysis_df, start_row, grand_label="Grand total")


def _append_emp_agency_total_block(ws, emp_agency_df: pd.DataFrame, start_row: int) -> int:
    """Write EMP / AGENCY / TOTAL summary below category block (cols B and D–H). Returns one past last row."""
    if emp_agency_df.empty:
        return start_row
    r = start_row
    ws.cell(r, _CATEGORY_COL, "Category")
    for i, h in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(r, _CATEGORY_NUM_START + i, h)
    r += 1
    for _, row in emp_agency_df.iterrows():
        label = row["Category"]
        ws.cell(r, _CATEGORY_COL, "" if pd.isna(label) else str(label))
        for i, h in enumerate(_CATEGORY_BAND_COLS):
            v = row[h]
            ws.cell(r, _CATEGORY_NUM_START + i, 0.0 if pd.isna(v) else float(v))
        r += 1
    return r


def _apply_excel_two_decimal_format(workbook: Any) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = _EXCEL_NUM_FORMAT


def _select_export_columns(df: pd.DataFrame, columns: list[str] | None) -> pd.DataFrame:
    if df.empty or not columns:
        return df
    present = [c for c in columns if c in df.columns]
    return df[present]


def build_excel_bytes(
    result: PayrollResult,
    *,
    column_rename: dict[str, str] | None = None,
    employee_columns: list[str] | None = None,
) -> bytes:
    all_df = _select_export_columns(pd.DataFrame(result.rows), employee_columns)
    agency_df = _select_export_columns(pd.DataFrame(result.agency_rows), employee_columns)
    gazebo_df = _select_export_columns(pd.DataFrame(result.gazebo_rows), employee_columns)
    analysis_df = _build_analysis_dataframe(all_df)
    emp_agency_df = build_emp_agency_total_df(result)
    category_hr_df = build_category_summary_hr_df(analysis_df)
    over60_df = build_hours_over_60_df(all_df)

    if column_rename:
        all_df = all_df.rename(columns=column_rename)
        if not agency_df.empty:
            agency_df = agency_df.rename(columns=column_rename)
        if not gazebo_df.empty:
            gazebo_df = gazebo_df.rename(columns=column_rename)
        if not over60_df.empty:
            over60_df = over60_df.rename(columns=column_rename)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        all_df.to_excel(writer, sheet_name="All Data", index=False)
        agency_df.to_excel(writer, sheet_name="Agency Employee", index=False)
        gazebo_df.to_excel(writer, sheet_name="Gazebo Employee", index=False)
        analysis_df.to_excel(writer, sheet_name="Analysis", index=False)

        if not analysis_df.empty:
            ws_all = writer.book["All Data"]
            border_max_col = _CATEGORY_NUM_START + len(_CATEGORY_BAND_COLS) - 1
            r = int(ws_all.max_row) + 2
            tier_titles = _ALL_DATA_SECTION_TITLES

            r = _append_section_title(ws_all, r, tier_titles[0])
            table_start = r
            end_row = _append_category_breakdown_block(ws_all, analysis_df, r)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)
            r = end_row + 1

            grouped_df = _build_grouped_analysis_dataframe(analysis_df)
            r = _append_section_title(ws_all, r, tier_titles[1])
            table_start = r
            end_row = _append_hour_totals_block(ws_all, grouped_df, r, grand_label=None)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)
            r = end_row + 1

            r = _append_section_title(ws_all, r, tier_titles[2])
            table_start = r
            end_row = _append_emp_agency_total_block(ws_all, emp_agency_df, r)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)
            r = end_row + 1

            overall_df = _build_overall_analysis_dataframe(analysis_df)
            r = _append_section_title(ws_all, r, tier_titles[3])
            table_start = r
            end_row = _append_hour_totals_block(ws_all, overall_df, r, grand_label="GRAND TOTAL")
            overall_totals = _band_totals_from_df(overall_df)
            emp_totals = _band_totals_from_emp_agency_total(emp_agency_df)
            diff_bands = {c: overall_totals[c] - emp_totals[c] for c in _CATEGORY_BAND_COLS}
            end_row = _append_band_values_row(ws_all, end_row, "Difference", diff_bands)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)

            ws_an = writer.book["Analysis"]
            an_start = int(ws_an.max_row) + 2
            _append_grand_total_row_openpyxl(ws_an, analysis_df, an_start)

        emp_agency_df.to_excel(writer, sheet_name="EMP Agency Total", index=False)
        category_hr_df.to_excel(writer, sheet_name="Category summary", index=False)
        over60_df.to_excel(writer, sheet_name="Hours over 60", index=False)
        _apply_excel_two_decimal_format(writer.book)

    output.seek(0)
    return output.getvalue()
