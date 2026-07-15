"""Branded export helpers for the daily report (CSV / PDF / branded Excel)."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


BRAND_NAME = "Gazebo HR"
BRAND_TAGLINE = "Day payroll & reporting"
PRIMARY_BLUE = "#003078"
ACCENT_BLUE = "#1d70b8"
COVER_STATS_GREY = "#D9D9D9"
COVER_BORDER_GREY = "#BFBFBF"
COVER_TEXT_DARK = "#333333"
COVER_TEXT_MUTED = "#666666"
DAY_REPORT_TITLE = "Day Report"
WEEK_REPORT_TITLE = "Week Report"
MONTH_REPORT_TITLE = "Month Report"
# GOV.UK typeface: GDS Transport is licensed for service.gov.uk only; Arial/Helvetica
# is the official fallback (see design-system.service.gov.uk/styles/typeface/).
UK_GOV_FONT = "Arial"
UK_GOV_FONT_PDF = "Helvetica"
UK_GOV_FONT_PDF_BOLD = "Helvetica-Bold"

EXPORT_COLUMNS = [
    "Name",
    "Category",
    "SageNo",
    "BasicHours",
    "MonFriOvertime",
    "SatSunOvertime",
    "AnnualHoliday",
    "TotalPaidHours",
    "ContractedHours",
    "Overtime",
]

WEEKLY_BRAND_TAGLINE = "Week payroll & reporting"

WEEKLY_EXPORT_COLUMNS = [
    "Name",
    "Category",
    "SageNo",
    "BasicHours",
    "MonFriOvertime",
    "SatSunOvertime",
    "AnnualHoliday",
    "TotalPaidHours",
    "ContractedHours",
    "ExtraHours",
    "AdditionalHolidayPay",
    "ContractHourMatch",
    "ContractMatchReason",
]

WEEKLY_EXPORT_HEADER_LABELS = {
    "TotalPaidHours": "Actual hours",
    "ContractedHours": "Contracted hours",
    "ExtraHours": "Extra hours",
    "AdditionalHolidayPay": "Additional Holiday pay",
}

_PDF_HEADER_SHORT_LABELS = {
    "MonFriOvertime": "Mon-Fri OT",
    "SatSunOvertime": "Sat-Sun OT",
    "AnnualHoliday": "Annual hol",
    "TotalPaidHours": "Total paid",
    "ContractedHours": "Contracted",
    "ExtraHours": "Extra hrs",
    "AdditionalHolidayPay": "Add. hol pay",
}

_PDF_COL_WEIGHTS: dict[str, float] = {
    "Name": 3.0,
    "Category": 2.4,
    "SageNo": 0.7,
}
_PDF_DEFAULT_COL_WEIGHT = 0.82
_PDF_WRAP_COLS = frozenset({"Name", "Category"})
_PDF_TABLE_FONT_SIZE = 7
_PDF_MARGIN_MM = 10


def weekly_export_header_labels(column_keys: list[str] | None = None) -> list[str]:
    keys = column_keys or WEEKLY_EXPORT_COLUMNS
    return [WEEKLY_EXPORT_HEADER_LABELS.get(k, k) for k in keys]


def _now_label() -> str:
    return datetime.now().strftime("%d %B %Y, %H:%M")


def _logo_path() -> Path | None:
    here = Path(__file__).resolve().parent
    candidate = here / "static" / "images" / "logo.png"
    return candidate if candidate.exists() else None


def _format_summary_hours(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.2f}"
    except (TypeError, ValueError):
        return str(value)


def _staff_summary_rows(summary: dict[str, Any]) -> list[tuple[str, str, bool]]:
    """Label, value, and whether to show the worked-count footnote."""
    rows: list[tuple[str, str, bool]] = []
    proc_date = summary.get("processing_date")
    if proc_date:
        rows.append(("Data Processing for date", str(proc_date), False))
    if "total_staff" in summary:
        rows.append(("Total staff count", str(summary.get("total_staff", "")), True))
        rows.extend([
            ("Agency staff", str(summary.get("agency_staff", "")), False),
            ("Gazebo staff", str(summary.get("gazebo_staff", "")), False),
            ("Gazebo staff on Paid holiday", str(summary.get("gazebo_paid_holiday", "")), False),
            ("Total paid hours", _format_summary_hours(summary.get("total_paid_hours")), False),
        ])
    return rows


def _report_metadata_rows(summary: dict[str, Any], *, report_title: str) -> list[tuple[str, str]]:
    return [
        ("Report name", report_title),
        ("Generated at", _now_label()),
        ("Operator", str(summary.get("operator") or "HR")),
    ]


def _write_csv_day_report_preamble(
    writer: Any,
    summary: dict[str, Any],
    *,
    report_title: str,
) -> None:
    writer.writerow([f"{BRAND_NAME} — {report_title}"])
    for label, value, show_footnote in _staff_summary_rows(summary):
        row = [label, value]
        if show_footnote:
            row.append("count only if they have worked")
        writer.writerow(row)
    writer.writerow([])
    writer.writerow(["Report metadata"])
    for label, value in _report_metadata_rows(summary, report_title=report_title):
        writer.writerow([label, value])


def _pdf_gov_paragraph(text: str, *, bold: bool = False, size: int = 9, color: str = COVER_TEXT_DARK) -> Paragraph:
    face = UK_GOV_FONT_PDF_BOLD if bold else UK_GOV_FONT_PDF
    weight = "<b>" if bold else ""
    weight_end = "</b>" if bold else ""
    return Paragraph(
        f"<font face='{face}' size='{size}' color='{color}'>{weight}{text}{weight_end}</font>",
        getSampleStyleSheet()["Normal"],
    )


def _append_pdf_day_report_summary(
    story: list[Any],
    summary: dict[str, Any],
    *,
    report_title: str,
) -> None:
    stats_data = [[label, value] for label, value, _ in _staff_summary_rows(summary)]
    if stats_data:
        stats_table = Table(stats_data, colWidths=[70 * mm, 40 * mm])
        stats_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(COVER_STATS_GREY)),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(COVER_TEXT_DARK)),
            ("FONTNAME", (0, 0), (0, -1), UK_GOV_FONT_PDF_BOLD),
            ("FONTNAME", (1, 0), (1, -1), UK_GOV_FONT_PDF),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(COVER_BORDER_GREY)),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(Spacer(1, 3 * mm))
        story.append(stats_table)
        for label, _, show_footnote in _staff_summary_rows(summary):
            if show_footnote:
                story.append(_pdf_gov_paragraph(
                    "count only if they have worked",
                    size=8,
                    color=COVER_TEXT_MUTED,
                ))
                break

    meta_data = [["Report metadata", ""]] + [
        [label, value] for label, value in _report_metadata_rows(summary, report_title=report_title)
    ]
    meta_table = Table(meta_data, colWidths=[45 * mm, 65 * mm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), UK_GOV_FONT_PDF),
        ("FONTNAME", (0, 0), (0, 0), UK_GOV_FONT_PDF_BOLD),
        ("FONTNAME", (0, 1), (0, -1), UK_GOV_FONT_PDF_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(COVER_TEXT_DARK)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("SPAN", (0, 0), (1, 0)),
    ]))
    story.append(Spacer(1, 4 * mm))
    story.append(meta_table)


def _pdf_col_widths(columns: list[str], total_width: float) -> list[float]:
    weights = [_PDF_COL_WEIGHTS.get(col, _PDF_DEFAULT_COL_WEIGHT) for col in columns]
    weight_sum = sum(weights)
    return [total_width * w / weight_sum for w in weights]


def _pdf_cell_style(*, bold: bool = False, size: int = _PDF_TABLE_FONT_SIZE) -> ParagraphStyle:
    return ParagraphStyle(
        name=f"PdfCell{'Bold' if bold else ''}{size}",
        fontName=UK_GOV_FONT_PDF_BOLD if bold else UK_GOV_FONT_PDF,
        fontSize=size,
        leading=size + 1,
        splitLongWords=True,
    )


def _pdf_paragraph(text: str, *, bold: bool = False, size: int = _PDF_TABLE_FONT_SIZE) -> Paragraph:
    weight_open = "<b>" if bold else ""
    weight_close = "</b>" if bold else ""
    safe = escape(str(text))
    return Paragraph(
        f"{weight_open}{safe}{weight_close}",
        _pdf_cell_style(bold=bold, size=size),
    )


def _pdf_header_label(column_key: str, display_headers: list[str], export_cols: list[str]) -> str:
    idx = export_cols.index(column_key) if column_key in export_cols else -1
    if idx >= 0 and idx < len(display_headers):
        return display_headers[idx]
    return _PDF_HEADER_SHORT_LABELS.get(column_key, column_key)


def _pdf_table_cell(value: Any, column_key: str) -> Any:
    text = _pdf_cell(value)
    if column_key == "Category":
        text = text.replace("_", "_\u200b")
    if column_key in _PDF_WRAP_COLS:
        return _pdf_paragraph(text)
    return text


def build_csv_bytes(
    rows: list[dict[str, Any]],
    summary: dict[str, Any] | None = None,
    *,
    report_title: str = DAY_REPORT_TITLE,
    columns: list[str] | None = None,
    column_headers: list[str] | None = None,
) -> bytes:
    export_cols = columns or EXPORT_COLUMNS
    headers = column_headers or export_cols
    buf = io.StringIO()
    writer = csv.writer(buf)
    if summary and "total_staff" in summary:
        _write_csv_day_report_preamble(writer, summary, report_title=report_title)
    else:
        writer.writerow([f"{BRAND_NAME} — {report_title}"])
        writer.writerow([f"Generated: {_now_label()}"])
        if summary:
            writer.writerow([
                f"Total rows: {summary.get('total_rows', '')}",
                f"Agency: {summary.get('agency_rows', '')}",
                f"Gazebo: {summary.get('gazebo_rows', '')}",
                f"Total paid hours: {summary.get('total_paid_hours', '')}",
            ])
    writer.writerow([])
    writer.writerow(headers)
    for r in rows:
        out = []
        for c in export_cols:
            v = r.get(c, "")
            if isinstance(v, float):
                out.append(f"{v:.2f}")
            else:
                out.append(v)
        writer.writerow(out)
    return buf.getvalue().encode("utf-8-sig")


def build_weekly_csv_bytes(rows: list[dict[str, Any]], summary: dict[str, Any] | None = None) -> bytes:
    return build_csv_bytes(
        rows,
        summary,
        report_title="Weekly report",
        columns=WEEKLY_EXPORT_COLUMNS,
        column_headers=weekly_export_header_labels(),
    )


def add_branding_cover_sheet(
    xlsx_bytes: bytes,
    summary: dict[str, Any] | None = None,
    *,
    report_title: str = DAY_REPORT_TITLE,
    tagline: str | None = None,
) -> bytes:
    """Insert a branded cover sheet at index 0 of an existing xlsx workbook."""
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.create_sheet(title="Overview", index=0)

    header_fill = PatternFill("solid", fgColor=PRIMARY_BLUE.lstrip("#"))
    stats_fill = PatternFill("solid", fgColor=COVER_STATS_GREY.lstrip("#"))
    thin_side = Side(style="thin", color=COVER_BORDER_GREY.lstrip("#"))
    stats_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    title_font = Font(name=UK_GOV_FONT, size=22, bold=True, color="FFFFFF")
    label_font = Font(name=UK_GOV_FONT, size=11, bold=True, color=COVER_TEXT_DARK.lstrip("#"))
    value_font = Font(name=UK_GOV_FONT, size=11, color=COVER_TEXT_DARK.lstrip("#"))
    footnote_font = Font(name=UK_GOV_FONT, size=9, italic=True, color=COVER_TEXT_MUTED.lstrip("#"))
    meta_heading_font = Font(name=UK_GOV_FONT, size=11, bold=True, color=COVER_TEXT_DARK.lstrip("#"))
    meta_label_font = Font(name=UK_GOV_FONT, size=10, bold=True, color=COVER_TEXT_DARK.lstrip("#"))
    meta_value_font = Font(name=UK_GOV_FONT, size=10, color=COVER_TEXT_DARK.lstrip("#"))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 36

    ws.merge_cells("A1:C3")
    cell = ws["A1"]
    cell.value = f"{BRAND_NAME}\n{tagline or BRAND_TAGLINE}"
    cell.fill = header_fill
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 26

    logo = _logo_path()
    if logo is not None:
        try:
            img = XLImage(str(logo))
            img.width = 110
            img.height = 34
            ws.add_image(img, "C1")
        except Exception:
            pass

    start_row = 5
    summary_rows = _staff_summary_rows(summary or {})
    for offset, (label, value, show_footnote) in enumerate(summary_rows):
        r = start_row + offset
        ws.row_dimensions[r].height = 20
        for col in (1, 2):
            c = ws.cell(r, col)
            c.fill = stats_fill
            c.border = stats_border
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(r, 1, label).font = label_font
        ws.cell(r, 2, value).font = value_font
        if show_footnote:
            note = ws.cell(r, 3, "count only if they have worked")
            note.font = footnote_font
            note.alignment = Alignment(horizontal="left", vertical="center")

    meta_start = start_row + len(summary_rows) + 2
    ws.cell(meta_start, 1, "Report metadata").font = meta_heading_font
    for offset, (label, value) in enumerate(_report_metadata_rows(summary or {}, report_title=report_title)):
        r = meta_start + 1 + offset
        ws.cell(r, 1, label).font = meta_label_font
        ws.cell(r, 2, value).font = meta_value_font
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")

    note_row = meta_start + 5
    note = ws.cell(note_row, 1, f"This workbook was generated by {BRAND_NAME}.")
    note.font = footnote_font
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def add_weekly_branding_cover_sheet(xlsx_bytes: bytes, summary: dict[str, Any] | None = None) -> bytes:
    return add_branding_cover_sheet(
        xlsx_bytes,
        summary,
        report_title="Weekly report",
        tagline=WEEKLY_BRAND_TAGLINE,
    )


def build_pdf_bytes(
    rows: list[dict[str, Any]],
    summary: dict[str, Any] | None = None,
    *,
    report_title: str = DAY_REPORT_TITLE,
    columns: list[str] | None = None,
    column_headers: list[str] | None = None,
) -> bytes:
    export_cols = columns or EXPORT_COLUMNS
    headers = column_headers or export_cols
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=_PDF_MARGIN_MM * mm,
        rightMargin=_PDF_MARGIN_MM * mm,
        topMargin=_PDF_MARGIN_MM * mm,
        bottomMargin=12 * mm,
        title=f"{BRAND_NAME} — {report_title}",
        author=BRAND_NAME,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []

    logo = _logo_path()
    header_cells: list[Any] = []
    if logo is not None:
        try:
            header_cells.append(Image(str(logo), width=32 * mm, height=10 * mm))
        except Exception:
            header_cells.append("")
    else:
        header_cells.append("")
    header_cells.append(
        Paragraph(
            f"<font face='{UK_GOV_FONT_PDF_BOLD}' size='16' color='#FFFFFF'><b>{BRAND_NAME}</b></font><br/>"
            f"<font face='{UK_GOV_FONT_PDF}' size='10' color='#DDE3EE'>{report_title}</font>",
            styles["Normal"],
        ),
    )
    header = Table([header_cells], colWidths=[34 * mm, doc.width - 34 * mm])
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(PRIMARY_BLUE)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(header)

    if summary and "total_staff" in summary:
        _append_pdf_day_report_summary(story, summary, report_title=report_title)
    elif summary:
        meta = (
            f"<font face='{UK_GOV_FONT_PDF}' size='9' color='{COVER_TEXT_DARK}'>"
            f"<b>Total rows:</b> {summary.get('total_rows', '')} &nbsp;&nbsp;"
            f"<b>Agency:</b> {summary.get('agency_rows', '')} &nbsp;&nbsp;"
            f"<b>Gazebo:</b> {summary.get('gazebo_rows', '')} &nbsp;&nbsp;"
            f"<b>Total paid hours:</b> {summary.get('total_paid_hours', '')}"
            f"</font>"
        )
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(meta, styles["Normal"]))

    story.append(Spacer(1, 6 * mm))

    head = []
    for col in export_cols:
        label = escape(_pdf_header_label(col, headers, export_cols))
        head.append(Paragraph(
            f"<b>{label}</b>",
            ParagraphStyle(
                name="PdfHeaderCell",
                fontName=UK_GOV_FONT_PDF_BOLD,
                fontSize=_PDF_TABLE_FONT_SIZE,
                leading=_PDF_TABLE_FONT_SIZE + 1,
                textColor=colors.white,
            ),
        ))
    data: list[list[Any]] = [head]
    for r in rows:
        data.append([_pdf_table_cell(r.get(c, ""), c) for c in export_cols])

    col_widths = _pdf_col_widths(export_cols, doc.width)
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table_style: list[tuple[Any, ...]] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
        ("FONTNAME", (0, 1), (-1, -1), UK_GOV_FONT_PDF),
        ("FONTSIZE", (0, 0), (-1, -1), _PDF_TABLE_FONT_SIZE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]
    _PDF_RIGHT_ALIGN_COLS = frozenset({
        "BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday",
        "TotalPaidHours", "ContractedHours", "Overtime", "ExtraHours", "AdditionalHolidayPay",
    })
    for idx, col in enumerate(export_cols):
        align = "RIGHT" if col in _PDF_RIGHT_ALIGN_COLS else "LEFT"
        table_style.append(("ALIGN", (idx, 0), (idx, -1), align))
    table.setStyle(TableStyle(table_style))
    story.append(table)

    def _draw_footer(canvas: Any, doc_obj: Any) -> None:
        canvas.saveState()
        canvas.setFont(UK_GOV_FONT_PDF, 8)
        canvas.setFillColor(colors.HexColor(COVER_TEXT_MUTED))
        footer = f"{BRAND_NAME} • Generated {_now_label()} • Page {doc_obj.page}"
        canvas.drawCentredString(doc_obj.pagesize[0] / 2.0, 8 * mm, footer)
        canvas.restoreState()

    doc.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    return buf.getvalue()


def build_weekly_pdf_bytes(rows: list[dict[str, Any]], summary: dict[str, Any] | None = None) -> bytes:
    return build_pdf_bytes(
        rows,
        summary,
        report_title="Weekly report",
        columns=WEEKLY_EXPORT_COLUMNS,
        column_headers=weekly_export_header_labels(),
    )


def _pdf_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def export_filename(prefix: str, ext: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    return f"{prefix}_{stamp}.{ext}"


def format_report_date_label(raw: str | None) -> str:
    text = (raw or "").strip().removeprefix("D ").strip()
    if not text:
        return datetime.now().strftime("%d.%m.%Y")
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text.split()[0], fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return text


def branded_report_filename(report_title: str, ext: str, date_label: str | None = None) -> str:
    return f"{report_title} - {format_report_date_label(date_label)} - {BRAND_NAME}.{ext}"


def day_report_filename(ext: str, processing_date: str | None = None) -> str:
    return branded_report_filename(DAY_REPORT_TITLE, ext, processing_date)


def week_report_filename(ext: str, processing_date: str | None = None) -> str:
    return branded_report_filename(WEEK_REPORT_TITLE, ext, processing_date)


def month_report_filename(ext: str, date_label: str | None = None) -> str:
    return branded_report_filename(MONTH_REPORT_TITLE, ext, date_label)


def month_report_date_label(summaries: Iterable[Any]) -> str | None:
    last = None
    for summary in summaries:
        last = summary
    if last is None:
        return None
    end = getattr(last, "end_date", "") or getattr(last, "start_date", "")
    text = (end or "").strip()
    return text or None
