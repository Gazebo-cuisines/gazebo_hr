from __future__ import annotations

import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Border, Side


logger = logging.getLogger(__name__)

_TABLE_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ClockRite "Paid Hours (Inc Absence) Summary": Pay ID column B (0-based 1); annual
# hours duplicated at Excel H and L (0-based 7 and 11). See payroll fix plan.
_CLOCKRITE_PAY_ID_COL = 1
_CLOCKRITE_SAGE_HEADER_COL = 3
_CLOCKRITE_ANNUAL_H_COL = 7
_CLOCKRITE_ANNUAL_L_COL = 11

# Appended category breakdown / grand total (matches monthly export layout).
_CATEGORY_COL = 2
_CATEGORY_NUM_START = 4
_CATEGORY_BAND_COLS = ["BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"]
_HOUR_DECIMAL_FIELDS = frozenset(
    {
        "BasicHours",
        "MonFriOvertime",
        "SatSunOvertime",
        "AnnualHoliday",
        "TotalPaidHours",
        "ContractedHours",
        "Overtime",
        "ExtraHours",
        "AdditionalHolidayPay",
    }
)
_EXCEL_NUM_FORMAT = "0.00"

def is_agency_category(category: str) -> bool:
    """Agency rows use ClockRite categories whose name starts with A-."""
    return str(category or "").strip().upper().startswith("A-")


def agency_categories_from_rows(rows: list[dict[str, Any]]) -> list[str]:
    """Distinct category labels from rows that count as agency (A- prefix), sorted."""
    seen: set[str] = set()
    out: list[str] = []
    for row in rows:
        cat = str(row.get("Category", "")).strip()
        if is_agency_category(cat) and cat not in seen:
            seen.add(cat)
            out.append(cat)
    return sorted(out, key=str.upper)


def split_emp_agency_rows(
    employee_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (gazebo_rows, agency_rows) for EMP / Agency totals."""
    agency_rows = [r for r in employee_rows if is_agency_category(str(r.get("Category", "")))]
    gazebo_rows = [r for r in employee_rows if not is_agency_category(str(r.get("Category", "")))]
    return gazebo_rows, agency_rows


@dataclass
class PayrollResult:
    rows: list[dict[str, Any]]
    agency_rows: list[dict[str, Any]]
    gazebo_rows: list[dict[str, Any]]
    total_paid_hours: float


@dataclass(frozen=True)
class EmployeeContractBlock:
    payroll_number: int | None
    sage_pay_ref: int | None
    prox_id: int | None
    title_id: int | None
    full_name: str
    clock_name: str
    contract_hours: float
    source_row: int


@dataclass
class ContractFileIndex:
    blocks: list[EmployeeContractBlock]
    by_payroll: dict[int, float]
    by_name: dict[str, float]
    by_sage_name: dict[int, str]
    by_clock_name: dict[str, str]
    pay_id_to_block: dict[int, EmployeeContractBlock]
    name_to_block: dict[str, EmployeeContractBlock]
    conflicted_pay_ids: frozenset[int]
    conflicts: list[dict[str, Any]]


@dataclass
class ContractAuditResult:
    missing: list[dict[str, Any]]
    conflicts: list[dict[str, Any]]
    review: list[dict[str, Any]]


def total_paid_hours_from_rows(rows: list[dict[str, Any]]) -> float:
    """Sum TotalPaidHours across all employee rows (includes every category, including A- prefix)."""
    return round(sum(float(r.get("TotalPaidHours", 0.0)) for r in rows), 2)


def _normalize_header(text: str) -> str:
    return "".join(ch.lower() for ch in (text or "") if ch.isalnum())


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _parse_decimal(value: Any) -> float:
    text = _to_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _round2(value: float) -> float:
    return round(float(value), 2)


def _round_row_hours(row: dict[str, Any]) -> None:
    for key in _HOUR_DECIMAL_FIELDS:
        if key in row and row[key] is not None:
            row[key] = _round2(row[key])


def _parse_int(value: Any) -> int | None:
    text = _to_text(value).replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _row_contains_date_range(row: list[str]) -> bool:
    return any("date range" in cell.lower() for cell in row)


def _is_clockrite_paid_hours_summary_header(header_cells: list[str], pay_id_col: int) -> bool:
    """True when header row matches Paid Hours (Inc Absence) grid (Pay ID in B, Sage in D)."""
    if pay_id_col != _CLOCKRITE_PAY_ID_COL:
        return False
    if len(header_cells) <= _CLOCKRITE_SAGE_HEADER_COL:
        return False
    return _normalize_header(header_cells[_CLOCKRITE_SAGE_HEADER_COL]) == "sage"


def _annual_holiday_clockrite_hl(row: list[str]) -> float:
    """Annual leave from Excel columns H and L; same value — use H, then L if H empty."""
    h = _parse_decimal(row[_CLOCKRITE_ANNUAL_H_COL] if len(row) > _CLOCKRITE_ANNUAL_H_COL else "")
    l = _parse_decimal(row[_CLOCKRITE_ANNUAL_L_COL] if len(row) > _CLOCKRITE_ANNUAL_L_COL else "")
    if h and l and abs(h - l) > 0.001:
        return h
    if h:
        return h
    return l


def _is_category_row(row: list[str], pay_id_col: int) -> bool:
    pay = row[pay_id_col] if pay_id_col < len(row) else ""
    if _parse_int(pay) is not None:
        return False
    c0 = row[0] if row else ""
    return bool(c0 and not c0.lower().startswith("total for"))


def _load_sheet(file_obj: Any) -> pd.DataFrame:
    file_obj.seek(0)
    return pd.read_excel(file_obj, sheet_name=0, header=None, dtype=str)


def parse_employee_hours(file_obj: Any) -> list[dict[str, Any]]:
    df = _load_sheet(file_obj)
    rows = [[_to_text(v) for v in rec] for rec in df.values.tolist()]
    if not rows:
        return []

    header_row = -1
    pay_id_col = -1
    scan_limit = min(len(rows), 40)
    for r in range(scan_limit):
        for c, cell in enumerate(rows[r]):
            if _normalize_header(cell) == "payid":
                header_row = r
                pay_id_col = c
                break
        if header_row >= 0:
            break

    if header_row < 0 or pay_id_col < 0:
        return _parse_employee_legacy(rows)

    header_cells = rows[header_row]
    clockrite_grid = _is_clockrite_paid_hours_summary_header(header_cells, pay_id_col)

    name_col = pay_id_col - 1 if pay_id_col > 0 else 0
    basic_col = pay_id_col + 1
    if clockrite_grid:
        # Employee totals: basic C, MF OT D, SS OT E (header may label D/E as Sage / Hrs @ 1 / @ 2).
        mon_fri_col = pay_id_col + 2
        sat_sun_col = pay_id_col + 3
        annual_col = -1
    else:
        mon_fri_col = pay_id_col + 2
        sat_sun_col = pay_id_col + 3
        annual_col = pay_id_col + 4

    category = ""
    result: list[dict[str, Any]] = []
    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        if _row_contains_date_range(row):
            break
        first = row[0] if row else ""
        if first.lower().startswith("total for"):
            continue

        pay_text = row[pay_id_col] if pay_id_col < len(row) else ""
        name_text = row[name_col] if name_col < len(row) else ""
        pay_id = _parse_int(pay_text)
        if pay_id is None:
            if _is_category_row(row, pay_id_col):
                category = first
            continue

        if not name_text or name_text.upper() == "U EMPLOYEE" or "total for" in name_text.lower():
            continue

        basic_gross = _parse_decimal(row[basic_col] if basic_col < len(row) else "")
        mon_fri_ot = _parse_decimal(row[mon_fri_col] if mon_fri_col < len(row) else "")
        sat_sun_ot = _parse_decimal(row[sat_sun_col] if sat_sun_col < len(row) else "")
        if clockrite_grid:
            annual_holiday = _annual_holiday_clockrite_hl(row)
        else:
            annual_holiday = _parse_decimal(row[annual_col] if annual_col >= 0 and annual_col < len(row) else "")
        # Gross basic in export includes annual leave; net basic matches legacy path and FORMULAS #2.
        basic_hours = basic_gross - annual_holiday
        total_paid = basic_hours + mon_fri_ot + sat_sun_ot + annual_holiday

        row = {
            "Name": name_text.upper(),
            "Category": category,
            "SageNo": pay_id,
            "BasicHours": basic_hours,
            "MonFriOvertime": mon_fri_ot,
            "SatSunOvertime": sat_sun_ot,
            "AnnualHoliday": annual_holiday,
            "TotalPaidHours": total_paid,
        }
        _round_row_hours(row)
        result.append(row)

    return result


def _parse_employee_legacy(rows: list[list[str]]) -> list[dict[str, Any]]:
    category = ""
    expects_category = True
    result: list[dict[str, Any]] = []
    idx = 0
    while idx + 6 < len(rows):
        row = rows[6 + idx]
        col_a = row[0] if row else ""
        if col_a:
            if expects_category:
                if col_a.lower() != "default":
                    category = col_a
                expects_category = False
            else:
                if any("date range" in cell.lower() for cell in row):
                    break
                name = _to_text(row[0]).upper()
                if name and name != "U EMPLOYEE":
                    basic = _parse_decimal(row[2] if len(row) > 2 else "")
                    mon_fri = _parse_decimal(row[3] if len(row) > 3 else "")
                    sat_sun = _parse_decimal(row[6] if len(row) > 6 else "")
                    annual = _parse_decimal(row[7] if len(row) > 7 else "")
                    basic_after_legacy = basic - annual
                    result.append(
                        {
                            "Name": name,
                            "Category": category,
                            "SageNo": _parse_int(row[1] if len(row) > 1 else "") or 0,
                            "BasicHours": basic_after_legacy,
                            "MonFriOvertime": mon_fri,
                            "SatSunOvertime": sat_sun,
                            "AnnualHoliday": annual,
                            "TotalPaidHours": basic_after_legacy + mon_fri + sat_sun + annual,
                        }
                    )
        else:
            expects_category = True
            idx += 1
        idx += 1
    return result


def _find_header_index(headers: list[str], *candidates: str) -> int:
    for idx, cell in enumerate(headers):
        n = _normalize_header(cell)
        if any(n == c for c in candidates):
            return idx
    return -1


def _contract_hrs_value_in_row(row: list[str]) -> float | None:
    for c, cell in enumerate(row):
        n = _normalize_header(cell)
        if n in ("contracthrs", "contracthours", "contracthr", "contractedhours"):
            if c + 1 < len(row):
                return _parse_decimal(row[c + 1])
            return 0.0
    return None


def _block_start_row(rows: list[list[str]], r_payroll: int) -> int:
    """Row index of the block title row (Prox ID + full name) above Payroll Number."""
    for back in range(1, 30):
        if r_payroll - back < 0:
            break
        pr = rows[r_payroll - back]
        if not pr:
            continue
        a0 = _to_text(pr[0])
        if a0.isdigit() and len(a0) < 6 and len(pr) > 1 and _to_text(pr[1]):
            return r_payroll - back
    return max(0, r_payroll - 25)


def _parse_single_clockrite_block(rows: list[list[str]], r_payroll: int, pay_no: int) -> EmployeeContractBlock:
    block_start = _block_start_row(rows, r_payroll)
    hours = 0.0
    for back in range(1, r_payroll - block_start + 1):
        got = _contract_hrs_value_in_row(rows[r_payroll - back])
        if got is not None:
            hours = got
            break

    sage_pay_ref: int | None = None
    prox_id: int | None = None
    r1 = r_payroll + 1
    for rr in range(block_start, r1):
        row = rows[rr]
        for c, cell in enumerate(row):
            n = _normalize_header(cell)
            if n == "sagepayref" and c + 1 < len(row):
                sid = _parse_int(row[c + 1])
                if sid is not None and sid > 0:
                    sage_pay_ref = sid
            elif n == "proxid" and c + 1 < len(row):
                pid = _parse_int(row[c + 1])
                if pid is not None and pid > 0:
                    prox_id = pid

    full_name, clock_name = _block_names_near(rows, r_payroll)
    title_id = _parse_int(rows[block_start][0] if block_start < len(rows) and rows[block_start] else "")
    return EmployeeContractBlock(
        payroll_number=pay_no,
        sage_pay_ref=sage_pay_ref,
        prox_id=prox_id,
        title_id=title_id,
        full_name=full_name,
        clock_name=clock_name,
        contract_hours=hours,
        source_row=r_payroll + 1,
    )


def _parse_clockrite_blocks(rows: list[list[str]]) -> list[EmployeeContractBlock]:
    blocks: list[EmployeeContractBlock] = []
    for r, row in enumerate(rows):
        for c, cell in enumerate(row):
            if _normalize_header(cell) != "payrollnumber":
                continue
            if c + 1 >= len(row):
                continue
            pay_no = _parse_int(row[c + 1])
            if pay_no is None or pay_no == 0:
                continue
            blocks.append(_parse_single_clockrite_block(rows, r, pay_no))
    return blocks


def _block_pay_ids(block: EmployeeContractBlock) -> list[int]:
    ids: list[int] = []
    for value in (block.payroll_number, block.sage_pay_ref, block.prox_id, block.title_id):
        if value is not None and value > 0:
            ids.append(value)
    return ids


def _build_contract_index_from_blocks(blocks: list[EmployeeContractBlock]) -> ContractFileIndex:
    by_payroll: dict[int, float] = {}
    by_name: dict[str, float] = {}
    by_sage_name: dict[int, str] = {}
    by_clock_name: dict[str, str] = {}
    pay_id_to_block: dict[int, EmployeeContractBlock] = {}
    name_to_block: dict[str, EmployeeContractBlock] = {}
    conflicts: list[dict[str, Any]] = []
    conflicted_pay_ids: set[int] = set()

    def register_pay_id(pay_id: int, block: EmployeeContractBlock) -> None:
        if pay_id in conflicted_pay_ids:
            return
        existing = pay_id_to_block.get(pay_id)
        if existing is None:
            pay_id_to_block[pay_id] = block
            by_payroll[pay_id] = block.contract_hours
            return
        if abs(existing.contract_hours - block.contract_hours) > 0.001:
            conflicted_pay_ids.add(pay_id)
            conflicts.append(
                {
                    "PayId": pay_id,
                    "HoursA": existing.contract_hours,
                    "HoursB": block.contract_hours,
                    "RowA": existing.source_row,
                    "RowB": block.source_row,
                    "NameA": existing.full_name,
                    "NameB": block.full_name,
                }
            )

    def register_name(name: str, block: EmployeeContractBlock) -> None:
        key = name.upper()
        if not key:
            return
        existing = name_to_block.get(key)
        if existing is None:
            name_to_block[key] = block
            by_name[key] = block.contract_hours
            return
        if abs(existing.contract_hours - block.contract_hours) > 0.001:
            conflicts.append(
                {
                    "Name": name,
                    "HoursA": existing.contract_hours,
                    "HoursB": block.contract_hours,
                    "RowA": existing.source_row,
                    "RowB": block.source_row,
                }
            )

    for block in blocks:
        for pay_id in _block_pay_ids(block):
            register_pay_id(pay_id, block)
            if block.full_name and pay_id not in conflicted_pay_ids:
                by_sage_name[pay_id] = block.full_name
        if block.full_name:
            register_name(block.full_name, block)
        if block.clock_name:
            register_name(block.clock_name, block)
            by_clock_name[block.clock_name.upper()] = block.full_name

    return ContractFileIndex(
        blocks=blocks,
        by_payroll=by_payroll,
        by_name=by_name,
        by_sage_name=by_sage_name,
        by_clock_name=by_clock_name,
        pay_id_to_block=pay_id_to_block,
        name_to_block=name_to_block,
        conflicted_pay_ids=frozenset(conflicted_pay_ids),
        conflicts=conflicts,
    )


def _build_contract_index_from_tabular_rows(
    rows: list[list[str]], header_row: int, payroll_col: int, contract_col: int, name_col: int
) -> ContractFileIndex:
    blocks: list[EmployeeContractBlock] = []
    for offset, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        if _row_contains_date_range(row):
            break
        hours = _parse_decimal(row[contract_col] if contract_col < len(row) else "")
        if hours == 0.0:
            continue
        pay_no = _parse_int(row[payroll_col] if payroll_col < len(row) else "")
        name = _to_text(row[name_col] if name_col >= 0 and name_col < len(row) else "")
        blocks.append(
            EmployeeContractBlock(
                payroll_number=pay_no,
                sage_pay_ref=None,
                prox_id=None,
                title_id=None,
                full_name=name,
                clock_name=name,
                contract_hours=hours,
                source_row=offset,
            )
        )
    return _build_contract_index_from_blocks(blocks)


def load_contract_file_index(file_obj: Any) -> ContractFileIndex:
    df = _load_sheet(file_obj)
    rows = [[_to_text(v) for v in rec] for rec in df.values.tolist()]
    if not rows:
        return ContractFileIndex([], {}, {}, {}, {}, {}, {}, frozenset(), [])

    header_row = -1
    payroll_col = -1
    contract_col = -1
    name_col = -1
    for r, row in enumerate(rows[:80]):
        pc = _find_header_index(row, "payrollnumber", "payrollno", "payroll")
        cc = _find_header_index(row, "contracthrs", "contracthours", "contracthr", "contractedhours")
        if pc >= 0 and cc >= 0:
            header_row = r
            payroll_col = pc
            contract_col = cc
            name_col = _find_header_index(row, "clockname", "name", "employeename")
            break

    if header_row >= 0:
        index = _build_contract_index_from_tabular_rows(rows, header_row, payroll_col, contract_col, name_col)
        if index.blocks:
            return index

    return _build_contract_index_from_blocks(_parse_clockrite_blocks(rows))


def _block_names_near(rows: list[list[str]], r_payroll: int) -> tuple[str, str]:
    """Full name from block title row; clock name from Clock Name row above Payroll Number."""
    full_name = ""
    clock_name = ""
    for back in range(1, 30):
        if r_payroll - back < 0:
            break
        pr = rows[r_payroll - back]
        if not pr:
            continue
        for c, cell in enumerate(pr):
            if _normalize_header(cell) == "clockname" and c + 1 < len(pr):
                cn = _to_text(pr[c + 1])
                if cn:
                    clock_name = cn
        if full_name:
            continue
        a0 = _to_text(pr[0])
        if a0.isdigit() and len(a0) < 6 and len(pr) > 1:
            candidate = _to_text(pr[1])
            if candidate:
                full_name = candidate
    return full_name, clock_name


def _parse_clockrite_display_names(rows: list[list[str]]) -> tuple[dict[int, str], dict[str, str]]:
    index = _build_contract_index_from_blocks(_parse_clockrite_blocks(rows))
    return index.by_sage_name, index.by_clock_name


def parse_employee_display_names(file_obj: Any) -> tuple[dict[int, str], dict[str, str]]:
    """Full names from contract file (ClockRite Employee Details layout)."""
    index = load_contract_file_index(file_obj)
    return index.by_sage_name, index.by_clock_name


def _parse_clockrite_contract_report(rows: list[list[str]]) -> tuple[dict[int, float], dict[str, float]]:
    index = _build_contract_index_from_blocks(_parse_clockrite_blocks(rows))
    return index.by_payroll, index.by_name


def parse_contracted_hours(file_obj: Any) -> tuple[dict[int, float], dict[str, float]]:
    index = load_contract_file_index(file_obj)
    return index.by_payroll, index.by_name


def audit_contract_pay_id_coverage(
    employee_rows: list[dict[str, Any]], contracted_file_obj: Any
) -> list[dict[str, Any]]:
    """Employees whose Pay ID does not appear in the contract export (Payroll Number / Sage Pay Ref)."""
    contracted_file_obj.seek(0)
    index = load_contract_file_index(contracted_file_obj)
    missing: list[dict[str, Any]] = []
    for row in employee_rows:
        sage_no = int(row["SageNo"])
        if sage_no not in index.by_payroll:
            missing.append(
                {
                    "SageNo": sage_no,
                    "Name": row.get("Name"),
                    "Category": row.get("Category"),
                }
            )
    return missing


def _same_person_blocks(
    block_a: EmployeeContractBlock,
    block_b: EmployeeContractBlock,
    sage_no: int,
    name_upper: str,
    index: ContractFileIndex,
) -> bool:
    if block_a.source_row == block_b.source_row:
        return True
    if block_a.full_name and block_b.full_name and block_a.full_name.upper() == block_b.full_name.upper():
        return True
    sage_name = index.by_sage_name.get(sage_no, "").upper()
    if sage_name and block_b.full_name.upper() == sage_name:
        return True
    if name_upper and block_b.full_name.upper() == name_upper:
        return True
    if name_upper and block_b.clock_name.upper() == name_upper:
        return True
    return False


def _contract_candidates(
    sage_no: int,
    name_upper: str,
    index: ContractFileIndex,
) -> list[tuple[EmployeeContractBlock, str]]:
    candidates: list[tuple[EmployeeContractBlock, str]] = []
    seen_rows: set[int] = set()

    def add(block: EmployeeContractBlock | None, reason: str) -> None:
        if block is None or block.source_row in seen_rows:
            return
        seen_rows.add(block.source_row)
        candidates.append((block, reason))

    if sage_no in index.conflicted_pay_ids:
        return candidates

    add(index.pay_id_to_block.get(sage_no), "Pay ID")
    add(index.name_to_block.get(name_upper), "employee name")
    sage_name = index.by_sage_name.get(sage_no, "").upper()
    if sage_name:
        add(index.name_to_block.get(sage_name), "Sage name map")
    clock_full = index.by_clock_name.get(name_upper, "").upper()
    if clock_full:
        add(index.name_to_block.get(clock_full), "clock name lookup")
    return candidates


def _resolve_contracted_hours(
    sage_no: int,
    name_upper: str,
    index: ContractFileIndex,
) -> tuple[float, str, str]:
    """Return (contracted hours, ContractHourMatch, ContractMatchReason)."""
    if sage_no in index.conflicted_pay_ids:
        return 0.0, "Review", "Contract file ID conflict — manual review"

    candidates = _contract_candidates(sage_no, name_upper, index)
    if not candidates:
        return 0.0, "No", "Pay ID not in contract export"

    unique_blocks = {block.source_row: block for block, _ in candidates}
    unique_hours = {block.contract_hours for block in unique_blocks.values()}
    if len(unique_hours) > 1:
        return 0.0, "Review", "Contract file ID conflict — manual review"

    pay_block = index.pay_id_to_block.get(sage_no)
    if pay_block is not None:
        hours = pay_block.contract_hours
        if hours == 0.0:
            for block, reason in candidates:
                if block.contract_hours > 0.0 and _same_person_blocks(pay_block, block, sage_no, name_upper, index):
                    if reason == "Pay ID":
                        continue
                    return float(block.contract_hours), "Yes", f"Matched on {reason} (Pay ID had zero contract hours)"
        return float(hours), "Yes", "Matched on Pay ID"

    block, reason = candidates[0]
    if reason == "employee name":
        reason = "employee name"
    elif reason == "clock name lookup":
        reason = "employee name (clock name lookup)"
    return float(block.contract_hours), "Yes", f"Matched on {reason}"


def audit_contract_integrity(
    employee_rows: list[dict[str, Any]],
    contracted_file_obj: Any,
    payroll_rows: list[dict[str, Any]] | None = None,
) -> ContractAuditResult:
    contracted_file_obj.seek(0)
    index = load_contract_file_index(contracted_file_obj)
    missing: list[dict[str, Any]] = []
    for row in employee_rows:
        sage_no = int(row["SageNo"])
        if sage_no not in index.by_payroll:
            missing.append(
                {
                    "SageNo": sage_no,
                    "Name": row.get("Name"),
                    "Category": row.get("Category"),
                }
            )
    review: list[dict[str, Any]] = []
    if payroll_rows:
        for row in payroll_rows:
            if row.get("ContractHourMatch") == "Review":
                review.append(
                    {
                        "SageNo": row.get("SageNo"),
                        "Name": row.get("Name"),
                        "Category": row.get("Category"),
                        "ContractMatchReason": row.get("ContractMatchReason"),
                    }
                )
    return ContractAuditResult(missing=missing, conflicts=index.conflicts, review=review)


def calculate_payroll(employee_rows: list[dict[str, Any]], contracted_file_obj: Any) -> PayrollResult:
    contracted_file_obj.seek(0)
    index = load_contract_file_index(contracted_file_obj)

    for row in employee_rows:
        name_upper = str(row["Name"]).upper()
        contracted, match, reason = _resolve_contracted_hours(int(row["SageNo"]), name_upper, index)
        row["ContractHourMatch"] = match
        row["ContractMatchReason"] = reason
        if match == "No":
            logger.debug(
                "Contract match miss: SageNo=%s Name=%s (%s)",
                row.get("SageNo"),
                row.get("Name"),
                reason,
            )
        row["ContractedHours"] = _round2(contracted)
        row["Overtime"] = _round2(max(0.0, float(row["TotalPaidHours"]) - contracted))
        _round_row_hours(row)

    for row in employee_rows:
        full = index.by_sage_name.get(int(row["SageNo"])) or index.by_clock_name.get(str(row["Name"]).upper())
        if full:
            row["Name"] = full

    gazebo_rows, agency_rows = split_emp_agency_rows(employee_rows)
    all_hours = total_paid_hours_from_rows(employee_rows)

    return PayrollResult(
        rows=employee_rows,
        agency_rows=agency_rows,
        gazebo_rows=gazebo_rows,
        total_paid_hours=all_hours,
    )


_HOLIDAY_PAY_FACTOR = 0.1207
HOLIDAY_PAY_FACTOR = _HOLIDAY_PAY_FACTOR


def compute_extra_holiday_pay(
    total_paid_hours: float,
    contracted_hours: float = 0.0,
    *,
    extra_hours: float | None = None,
    additional_holiday_pay: float | None = None,
) -> tuple[float, float]:
    """Weekly report rules: extra = max(0, actual − contracted); holiday = factor × extra."""
    if extra_hours is not None:
        extra = _round2(max(0.0, float(extra_hours)))
    else:
        actual = max(0.0, float(total_paid_hours))
        contracted = max(0.0, float(contracted_hours))
        extra = _round2(max(0.0, actual - contracted))
    if additional_holiday_pay is not None:
        holiday = _round2(max(0.0, float(additional_holiday_pay)))
    else:
        holiday = _round2(max(0.0, extra * _HOLIDAY_PAY_FACTOR))
    return extra, holiday


def calculate_weekly_payroll(employee_rows: list[dict[str, Any]], contracted_file_obj: Any) -> PayrollResult:
    """Like calculate_payroll, with clamped hours, ExtraHours, and AdditionalHolidayPay."""
    result = calculate_payroll(employee_rows, contracted_file_obj)
    for row in result.rows:
        actual = max(0.0, float(row["TotalPaidHours"]))
        contracted = max(0.0, float(row["ContractedHours"]))
        row["TotalPaidHours"] = actual
        row["ContractedHours"] = contracted
        row.pop("Overtime", None)
        extra, holiday = compute_extra_holiday_pay(actual, contracted)
        row["ExtraHours"] = extra
        row["AdditionalHolidayPay"] = holiday
        _round_row_hours(row)
    return PayrollResult(
        rows=result.rows,
        agency_rows=result.agency_rows,
        gazebo_rows=result.gazebo_rows,
        total_paid_hours=total_paid_hours_from_rows(result.rows),
    )


_HOUR_BAND_COLS = ["BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"]


def _sum_hour_bands(rows: list[dict[str, Any]]) -> dict[str, float]:
    return {k: sum(float(r.get(k, 0.0) or 0.0) for r in rows) for k in _HOUR_BAND_COLS}


def build_emp_agency_total_df(result: PayrollResult) -> pd.DataFrame:
    """EMP (non-agency / Gazebo), AGENCY, and TOTAL sums for the five hour bands (section C)."""
    emp = _sum_hour_bands(result.gazebo_rows)
    agency = _sum_hour_bands(result.agency_rows)
    total = _sum_hour_bands(result.rows)
    return pd.DataFrame(
        [
            {"Category": "EMP", **emp},
            {"Category": "AGENCY", **agency},
            {"Category": "TOTAL", **total},
        ]
    )


def build_category_summary_hr_df(analysis_df: pd.DataFrame) -> pd.DataFrame:
    """Same data as Analysis with HR-friendly column names and a grand total row (section B)."""
    hr_cols = {
        "BasicHours": "Basic Hour",
        "MonFriOvertime": "Mon Fri O",
        "SatSunOvertime": "Sat Sun O",
        "AnnualHoliday": "Annual Ho",
        "TotalPaidHours": "Total Paid Hours",
    }
    if analysis_df.empty:
        return pd.DataFrame(columns=["Category", *list(hr_cols.values())])
    out = analysis_df.rename(columns=hr_cols)
    total_row: dict[str, Any] = {"Category": "Grand total"}
    for src, dst in hr_cols.items():
        total_row[dst] = float(analysis_df[src].sum())
    return pd.concat([out, pd.DataFrame([total_row])], ignore_index=True)


def build_hours_over_60_df(all_df: pd.DataFrame) -> pd.DataFrame:
    """Employees with TotalPaidHours strictly greater than 60."""
    if all_df.empty:
        return pd.DataFrame(columns=["Name", "Category", "SageNo", "TotalPaidHours", "BasicHours", "Overtime", "ContractedHours"])
    work = all_df.copy()
    work["_tp"] = pd.to_numeric(work["TotalPaidHours"], errors="coerce").fillna(0.0)
    filtered = work[work["_tp"] > 60.0].drop(columns=["_tp"], errors="ignore")
    cols = ["Name", "Category", "SageNo", "TotalPaidHours", "BasicHours", "Overtime", "ContractedHours"]
    have = [c for c in cols if c in filtered.columns]
    return filtered[have] if have else filtered


def _build_analysis_dataframe(all_df: pd.DataFrame) -> pd.DataFrame:
    if all_df.empty:
        return pd.DataFrame(columns=["Category", "BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"])
    return (
        all_df.groupby("Category", dropna=False)[["BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours"]]
        .sum()
        .reset_index()
    )


_OVERALL_CATEGORY_ORDER = ("PROD", "PACK", "WRHS", "CLNR", "TECH", "OFFICE")


def _overall_category_key(category: str) -> str:
    c = str(category).strip().upper()
    if "PKNG" in c or "DPCH" in c:
        return "PACK"
    if "PROD" in c:
        return "PROD"
    if "WRHS" in c:
        return "WRHS"
    if "CLNR" in c:
        return "CLNR"
    if "TECH" in c:
        return "TECH"
    if "OFCE" in c:
        return "OFFICE"
    return "OTHER"


def _build_grouped_analysis_dataframe(analysis_df: pd.DataFrame) -> pd.DataFrame:
    from .monthly_service import _grouped_key

    if analysis_df.empty:
        return pd.DataFrame(columns=["Category", *_CATEGORY_BAND_COLS])
    grouped: dict[str, dict[str, float]] = {}
    for _, row in analysis_df.iterrows():
        k = _grouped_key(str(row["Category"]))
        if k not in grouped:
            grouped[k] = {c: 0.0 for c in _CATEGORY_BAND_COLS}
        for c in _CATEGORY_BAND_COLS:
            grouped[k][c] += float(row[c] or 0.0)
    return pd.DataFrame([{"Category": k, **grouped[k]} for k in sorted(grouped)])


def _build_overall_analysis_dataframe(analysis_df: pd.DataFrame) -> pd.DataFrame:
    if analysis_df.empty:
        return pd.DataFrame(columns=["Category", *_CATEGORY_BAND_COLS])
    buckets = {k: {c: 0.0 for c in _CATEGORY_BAND_COLS} for k in _OVERALL_CATEGORY_ORDER}
    for _, row in analysis_df.iterrows():
        b = _overall_category_key(str(row["Category"]))
        if b not in buckets:
            buckets[b] = {c: 0.0 for c in _CATEGORY_BAND_COLS}
        for c in _CATEGORY_BAND_COLS:
            buckets[b][c] += float(row[c] or 0.0)
    return pd.DataFrame([{"Category": k, **buckets[k]} for k in _OVERALL_CATEGORY_ORDER])


def build_overall_category_totals(rows: list[dict[str, Any]]) -> pd.DataFrame:
    """Roll per-category band sums into PROD/PACK/WRHS/CLNR/TECH/OFFICE (weekly + monthly)."""
    if not rows:
        return pd.DataFrame(columns=["Category", *_CATEGORY_BAND_COLS])
    return _build_overall_analysis_dataframe(pd.DataFrame(rows))


_ALL_DATA_SECTION_TITLES = (
    "Category breakdown (detailed)",
    "Category breakdown (grouped)",
    "EMP / Agency totals",
    "Category breakdown (overall)",
)


def _append_section_title(ws, start_row: int, title: str) -> int:
    """Write a section label in col B. Returns the next row for the table header."""
    ws.cell(start_row, _CATEGORY_COL, title)
    return start_row + 1


def _apply_table_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Apply thin borders to every cell in the given range (formatting only)."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = _TABLE_BORDER


def _band_totals_from_df(df: pd.DataFrame) -> dict[str, float]:
    return {c: float(df[c].sum()) for c in _CATEGORY_BAND_COLS}


def _band_totals_from_emp_agency_total(emp_agency_df: pd.DataFrame) -> dict[str, float]:
    total_row = emp_agency_df.loc[emp_agency_df["Category"] == "TOTAL"].iloc[0]
    return {c: float(total_row[c]) for c in _CATEGORY_BAND_COLS}


def _append_band_values_row(ws, start_row: int, label: str, bands: dict[str, float]) -> int:
    """Write a labeled row of hour-band values (cols B and D–H). Returns one past the last row."""
    ws.cell(start_row, _CATEGORY_COL, label)
    for i, col in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(start_row, _CATEGORY_NUM_START + i, float(bands[col]))
    return start_row + 1


def _append_grand_total_row_openpyxl(
    ws,
    totals_df: pd.DataFrame,
    start_row: int,
    *,
    label: str = "Grand total",
) -> int:
    """Write total label and column sums. Returns one past the last row written."""
    if totals_df.empty:
        return start_row
    r = start_row
    ws.cell(r, _CATEGORY_COL, label)
    for i, col in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(r, _CATEGORY_NUM_START + i, float(totals_df[col].sum()))
    return r + 1


def _append_hour_totals_block(
    ws,
    totals_df: pd.DataFrame,
    start_row: int,
    *,
    header_label: str = "Category",
    grand_label: str | None = "Grand total",
) -> int:
    """Write category/hour band table (cols B and D–H). Returns one past the last row."""
    if totals_df.empty:
        return start_row
    r = start_row
    ws.cell(r, _CATEGORY_COL, header_label)
    for i, h in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(r, _CATEGORY_NUM_START + i, h)
    r += 1
    for _, row in totals_df.iterrows():
        cat = row["Category"]
        ws.cell(r, _CATEGORY_COL, "" if pd.isna(cat) else str(cat))
        for i, h in enumerate(_CATEGORY_BAND_COLS):
            v = row[h]
            ws.cell(r, _CATEGORY_NUM_START + i, 0.0 if pd.isna(v) else float(v))
        r += 1
    if grand_label:
        r += 1
        return _append_grand_total_row_openpyxl(ws, totals_df, r, label=grand_label)
    return r


def _append_category_breakdown_block(ws, analysis_df: pd.DataFrame, start_row: int) -> int:
    """Granular category breakdown with grand total (tier 1)."""
    return _append_hour_totals_block(ws, analysis_df, start_row, grand_label="Grand total")


def _append_emp_agency_total_block(ws, emp_agency_df: pd.DataFrame, start_row: int) -> int:
    """Write EMP / AGENCY / TOTAL summary below category block (cols B and D–H). Returns one past last row."""
    if emp_agency_df.empty:
        return start_row
    r = start_row
    ws.cell(r, _CATEGORY_COL, "Category")
    for i, h in enumerate(_CATEGORY_BAND_COLS):
        ws.cell(r, _CATEGORY_NUM_START + i, h)
    r += 1
    for _, row in emp_agency_df.iterrows():
        label = row["Category"]
        ws.cell(r, _CATEGORY_COL, "" if pd.isna(label) else str(label))
        for i, h in enumerate(_CATEGORY_BAND_COLS):
            v = row[h]
            ws.cell(r, _CATEGORY_NUM_START + i, 0.0 if pd.isna(v) else float(v))
        r += 1
    return r


def _apply_excel_two_decimal_format(workbook: Any) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = _EXCEL_NUM_FORMAT


def build_excel_bytes(result: PayrollResult, *, column_rename: dict[str, str] | None = None) -> bytes:
    all_df = pd.DataFrame(result.rows)
    agency_df = pd.DataFrame(result.agency_rows)
    gazebo_df = pd.DataFrame(result.gazebo_rows)
    analysis_df = _build_analysis_dataframe(all_df)
    emp_agency_df = build_emp_agency_total_df(result)
    category_hr_df = build_category_summary_hr_df(analysis_df)
    over60_df = build_hours_over_60_df(all_df)

    if column_rename:
        all_df = all_df.rename(columns=column_rename)
        if not agency_df.empty:
            agency_df = agency_df.rename(columns=column_rename)
        if not gazebo_df.empty:
            gazebo_df = gazebo_df.rename(columns=column_rename)
        if not over60_df.empty:
            over60_df = over60_df.rename(columns=column_rename)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        all_df.to_excel(writer, sheet_name="All Data", index=False)
        agency_df.to_excel(writer, sheet_name="Agency Employee", index=False)
        gazebo_df.to_excel(writer, sheet_name="Gazebo Employee", index=False)
        analysis_df.to_excel(writer, sheet_name="Analysis", index=False)

        if not analysis_df.empty:
            ws_all = writer.book["All Data"]
            border_max_col = _CATEGORY_NUM_START + len(_CATEGORY_BAND_COLS) - 1
            r = int(ws_all.max_row) + 2
            tier_titles = _ALL_DATA_SECTION_TITLES

            r = _append_section_title(ws_all, r, tier_titles[0])
            table_start = r
            end_row = _append_category_breakdown_block(ws_all, analysis_df, r)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)
            r = end_row + 1

            grouped_df = _build_grouped_analysis_dataframe(analysis_df)
            r = _append_section_title(ws_all, r, tier_titles[1])
            table_start = r
            end_row = _append_hour_totals_block(ws_all, grouped_df, r, grand_label=None)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)
            r = end_row + 1

            r = _append_section_title(ws_all, r, tier_titles[2])
            table_start = r
            end_row = _append_emp_agency_total_block(ws_all, emp_agency_df, r)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)
            r = end_row + 1

            overall_df = _build_overall_analysis_dataframe(analysis_df)
            r = _append_section_title(ws_all, r, tier_titles[3])
            table_start = r
            end_row = _append_hour_totals_block(ws_all, overall_df, r, grand_label="GRAND TOTAL")
            overall_totals = _band_totals_from_df(overall_df)
            emp_totals = _band_totals_from_emp_agency_total(emp_agency_df)
            diff_bands = {c: overall_totals[c] - emp_totals[c] for c in _CATEGORY_BAND_COLS}
            end_row = _append_band_values_row(ws_all, end_row, "Difference", diff_bands)
            _apply_table_border(ws_all, table_start, end_row - 1, _CATEGORY_COL, border_max_col)

            ws_an = writer.book["Analysis"]
            an_start = int(ws_an.max_row) + 2
            _append_grand_total_row_openpyxl(ws_an, analysis_df, an_start)

        emp_agency_df.to_excel(writer, sheet_name="EMP Agency Total", index=False)
        category_hr_df.to_excel(writer, sheet_name="Category summary", index=False)
        over60_df.to_excel(writer, sheet_name="Hours over 60", index=False)
        _apply_excel_two_decimal_format(writer.book)

    output.seek(0)
    return output.getvalue()
