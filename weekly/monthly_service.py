from __future__ import annotations

from dataclasses import asdict, dataclass, field
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .payroll_service import build_overall_category_totals, compute_extra_holiday_pay, is_agency_category


_BAND_KEYS = ("BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours")
_HOLIDAY_PAY_KEYS = ("ContractedHours", "ExtraHours", "AdditionalHolidayPay")
_EMPLOYEE_VALUE_KEYS = _BAND_KEYS + _HOLIDAY_PAY_KEYS
_EMP_AGENCY_ROWS = ("EMP", "AGENCY", "TOTAL")
_SHEET_LAST_COL = 11
_PAY_ID_COL = 3
_NUM_FORMAT = "0.00"
_INTEGER_FORMAT = "0"
_PRIMARY_BLUE = "003078"

_MONTHLY_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill("solid", fgColor=_PRIMARY_BLUE)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_SECTION_TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="333333")
_SECTION_SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
_SHEET_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=_PRIMARY_BLUE)

_BAND_LABELS = (
    "Basic hours",
    "Mon–Fri overtime",
    "Sat/Sun overtime",
    "Annual holiday",
    "Total paid hours",
)
_HOLIDAY_PAY_LABELS = (
    "Contracted hours",
    "Extra hours",
    "Additional holiday pay",
)


@dataclass
class MonthlyEmployee:
    Name: str
    Category: str
    SageNo: int
    BasicHours: float
    MonFriOvertime: float
    SatSunOvertime: float
    AnnualHoliday: float
    TotalPaidHours: float
    ContractedHours: float = 0.0
    ExtraHours: float = 0.0
    AdditionalHolidayPay: float = 0.0


@dataclass
class MonthlyEmployeeTotal:
    Category: str
    BasicHours: float
    MonFriOvertime: float
    SatSunOvertime: float
    AnnualHoliday: float
    TotalPaidHours: float


@dataclass
class MonthlyAdjustment:
    Name: str
    Type: str
    Value: float


@dataclass
class MonthlyWeekSummary:
    employees: list[MonthlyEmployee] = field(default_factory=list)
    employee_totals: list[MonthlyEmployeeTotal] = field(default_factory=list)
    adjustments: list[MonthlyAdjustment] = field(default_factory=list)
    start_date: str = ""
    end_date: str = ""
    non_agency_total: float = 0.0
    grouped_totals: dict[str, MonthlyEmployeeTotal] = field(default_factory=dict)
    emp_agency_bands: dict[str, dict[str, float]] = field(default_factory=dict)
    total_extra_hours: float = 0.0
    total_additional_holiday_pay: float = 0.0


@dataclass
class HolidayPayLayout:
    header_row: int
    rows: dict[str, int]
    next_row: int


@dataclass
class EmployeeTableLayout:
    header_row: int
    data_start: int
    data_end: int
    next_row: int


@dataclass
class EmpAgencyLayout:
    header_row: int
    rows: dict[str, int]
    next_row: int


@dataclass
class WeekSheetLayout:
    sheet_name: str
    employee: EmployeeTableLayout
    emp_agency: EmpAgencyLayout
    holiday_pay: HolidayPayLayout


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _parse_decimal(value: Any) -> float:
    text = _to_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_int(value: Any) -> int:
    text = _to_text(value).replace(",", "")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def _cell(row: list[str], index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    return row[index]


def _all_data_column_indices(header_row: list[str]) -> dict[str, int]:
    from .payroll_service import _normalize_header

    by_norm = {_normalize_header(cell): i for i, cell in enumerate(header_row)}

    def col(*names: str) -> int:
        for name in names:
            idx = by_norm.get(_normalize_header(name))
            if idx is not None:
                return idx
        return -1

    return {
        "name": col("name"),
        "category": col("category"),
        "sage": col("sageno", "pay id", "pay id (sage)"),
        "basic": col("basichours", "basic hours"),
        "mon_fri": col("monfriovertime", "mon fri overtime", "mon–fri overtime"),
        "sat_sun": col("satsunovertime", "sat sun overtime", "sat/sun overtime"),
        "annual": col("annualholiday", "annual holiday"),
        "total": col("totalpaidhours", "actual hours", "total paid hours"),
        "contracted": col("contractedhours", "contracted hours"),
        "extra": col("extrahours", "extra hours"),
        "holiday_pay": col("additionalholidaypay", "additional holiday pay"),
    }


def _monthly_employee_from_row(row: list[str], cols: dict[str, int]) -> MonthlyEmployee:
    def dec(key: str, fallback: int) -> float:
        idx = cols.get(key, fallback)
        return _parse_decimal(_cell(row, idx))

    contracted_idx = cols.get("contracted", -1)
    extra_idx = cols.get("extra", -1)
    holiday_idx = cols.get("holiday_pay", -1)
    contracted = dec("contracted", -1) if contracted_idx >= 0 else 0.0
    parsed_extra = dec("extra", -1) if extra_idx >= 0 else None
    parsed_holiday = dec("holiday_pay", -1) if holiday_idx >= 0 else None

    employee = MonthlyEmployee(
        Name=_to_text(_cell(row, cols.get("name", 0))),
        Category=_to_text(_cell(row, cols.get("category", 1))),
        SageNo=_parse_int(_cell(row, cols.get("sage", 2))),
        BasicHours=dec("basic", 3),
        MonFriOvertime=dec("mon_fri", 4),
        SatSunOvertime=dec("sat_sun", 5),
        AnnualHoliday=dec("annual", 6),
        TotalPaidHours=dec("total", 7),
        ContractedHours=contracted,
    )
    extra, holiday = compute_extra_holiday_pay(
        employee.TotalPaidHours,
        employee.ContractedHours,
        extra_hours=parsed_extra,
        additional_holiday_pay=parsed_holiday,
    )
    employee.ExtraHours = extra
    employee.AdditionalHolidayPay = holiday
    return employee


def _grouped_key(category: str) -> str:
    category = _to_text(category)
    if category.startswith("A-") and len(category) >= 9:
        return f"{category[:4]} {category[5:9]}"
    return category[:4] if len(category) >= 4 else category


def _empty_bands() -> dict[str, float]:
    return {k: 0.0 for k in _BAND_KEYS}


def _compute_emp_agency_bands(employees: list[MonthlyEmployee]) -> dict[str, dict[str, float]]:
    emp = _empty_bands()
    agency = _empty_bands()
    for e in employees:
        target = agency if is_agency_category(e.Category) else emp
        target["BasicHours"] += e.BasicHours
        target["MonFriOvertime"] += e.MonFriOvertime
        target["SatSunOvertime"] += e.SatSunOvertime
        target["AnnualHoliday"] += e.AnnualHoliday
        target["TotalPaidHours"] += e.TotalPaidHours
    total = {k: emp[k] + agency[k] for k in _BAND_KEYS}
    return {"EMP": emp, "AGENCY": agency, "TOTAL": total}


def _sum_emp_agency_bands(week_bands: list[dict[str, dict[str, float]]]) -> dict[str, dict[str, float]]:
    out = {row: _empty_bands() for row in _EMP_AGENCY_ROWS}
    for bands in week_bands:
        for row in _EMP_AGENCY_ROWS:
            for k in _BAND_KEYS:
                out[row][k] += float(bands.get(row, {}).get(k, 0.0) or 0.0)
    return out


def _employee_totals_from_employees(employees: list[MonthlyEmployee]) -> list[MonthlyEmployeeTotal]:
    by_cat: dict[str, MonthlyEmployeeTotal] = {}
    for e in employees:
        if e.Category not in by_cat:
            by_cat[e.Category] = MonthlyEmployeeTotal(e.Category, 0.0, 0.0, 0.0, 0.0, 0.0)
        t = by_cat[e.Category]
        t.BasicHours += e.BasicHours
        t.MonFriOvertime += e.MonFriOvertime
        t.SatSunOvertime += e.SatSunOvertime
        t.AnnualHoliday += e.AnnualHoliday
        t.TotalPaidHours += e.TotalPaidHours
    return list(by_cat.values())


def _build_grouped_totals(totals: list[MonthlyEmployeeTotal]) -> dict[str, MonthlyEmployeeTotal]:
    grouped: dict[str, MonthlyEmployeeTotal] = {}
    for t in totals:
        k = _grouped_key(t.Category)
        if k not in grouped:
            grouped[k] = MonthlyEmployeeTotal(k, 0.0, 0.0, 0.0, 0.0, 0.0)
        g = grouped[k]
        g.BasicHours += t.BasicHours
        g.MonFriOvertime += t.MonFriOvertime
        g.SatSunOvertime += t.SatSunOvertime
        g.AnnualHoliday += t.AnnualHoliday
        g.TotalPaidHours += t.TotalPaidHours
    return grouped


def _employee_totals_to_band_rows(totals: list[MonthlyEmployeeTotal]) -> list[dict[str, Any]]:
    return [
        {
            "Category": t.Category,
            "BasicHours": t.BasicHours,
            "MonFriOvertime": t.MonFriOvertime,
            "SatSunOvertime": t.SatSunOvertime,
            "AnnualHoliday": t.AnnualHoliday,
            "TotalPaidHours": t.TotalPaidHours,
        }
        for t in totals
    ]


def _build_overall_totals_from_employee_totals(
    totals: list[MonthlyEmployeeTotal],
) -> list[tuple[str, MonthlyEmployeeTotal]]:
    overall_df = build_overall_category_totals(_employee_totals_to_band_rows(totals))
    if overall_df.empty:
        return []
    out: list[tuple[str, MonthlyEmployeeTotal]] = []
    for _, row in overall_df.iterrows():
        out.append(
            (
                str(row["Category"]),
                MonthlyEmployeeTotal(
                    str(row["Category"]),
                    float(row["BasicHours"]),
                    float(row["MonFriOvertime"]),
                    float(row["SatSunOvertime"]),
                    float(row["AnnualHoliday"]),
                    float(row["TotalPaidHours"]),
                ),
            )
        )
    return out


def _enrich_week_summary(out: MonthlyWeekSummary) -> None:
    if not out.employee_totals and out.employees:
        out.employee_totals = _employee_totals_from_employees(out.employees)
    out.non_agency_total = sum(t.TotalPaidHours for t in out.employee_totals if not is_agency_category(t.Category))
    if not out.grouped_totals:
        out.grouped_totals = _build_grouped_totals(out.employee_totals)
    out.emp_agency_bands = _compute_emp_agency_bands(out.employees)
    out.total_extra_hours = round(sum(e.ExtraHours for e in out.employees), 2)
    out.total_additional_holiday_pay = round(sum(e.AdditionalHolidayPay for e in out.employees), 2)


def _workbook_has_all_data(file_obj: Any) -> bool:
    file_obj.seek(0)
    try:
        xl = pd.ExcelFile(file_obj)
        return "All Data" in xl.sheet_names
    except Exception:
        return False
    finally:
        file_obj.seek(0)


def parse_weekly_gazebo_all_data(
    file_obj: Any,
    *,
    start_date: str = "",
    end_date: str = "",
) -> MonthlyWeekSummary:
    """Parse weekly Gazebo export (.xlsx) — employees from the All Data sheet."""
    file_obj.seek(0)
    df = pd.read_excel(file_obj, sheet_name="All Data", header=None, dtype=str)
    text_rows = [[_to_text(v) for v in row] for row in df.values.tolist()]
    out = MonthlyWeekSummary(start_date=start_date, end_date=end_date)
    if not text_rows:
        return out

    header_row = -1
    for i, row in enumerate(text_rows[:5]):
        if _to_text(row[0] if len(row) > 0 else "").lower() == "name" and _to_text(row[1] if len(row) > 1 else "").lower() == "category":
            header_row = i
            break
    if header_row < 0:
        return out

    header_cells = text_rows[header_row]
    cols = _all_data_column_indices(header_cells)

    for row in text_rows[header_row + 1 :]:
        name = _to_text(_cell(row, cols.get("name", 0)))
        col_b = _to_text(_cell(row, cols.get("category", 1)))
        if not name:
            if col_b.lower() in ("category", "category breakdown (overall)"):
                break
            break
        if col_b.startswith("Category breakdown"):
            break
        out.employees.append(_monthly_employee_from_row(row, cols))
    _enrich_week_summary(out)
    return out


def parse_monthly_week_file(file_obj: Any) -> MonthlyWeekSummary:
    from .payroll_service import _load_sheet  # reuse weekly reader

    table = _load_sheet(file_obj)
    rows = table.values.tolist()
    text_rows = [[_to_text(v) for v in row] for row in rows]
    if not text_rows:
        return MonthlyWeekSummary()

    out = MonthlyWeekSummary()
    out.start_date = _to_text(text_rows[0][4] if len(text_rows[0]) > 4 else "").removeprefix("D ").strip()
    out.end_date = _to_text(text_rows[0][7] if len(text_rows[0]) > 7 else "").removeprefix("D ").strip()

    r = 3  # row 4 in excel
    while r < len(text_rows):
        row = text_rows[r]
        if not _to_text(row[0] if len(row) > 0 else ""):
            break
        out.employees.append(
            _monthly_employee_from_row(
                row,
                {
                    "name": 0,
                    "category": 1,
                    "sage": 2,
                    "basic": 3,
                    "mon_fri": 4,
                    "sat_sun": 5,
                    "annual": 6,
                    "total": 7,
                    "contracted": -1,
                    "extra": -1,
                    "holiday_pay": -1,
                },
            )
        )
        r += 1

    adjustments_header = -1
    for i in range(r, min(len(text_rows), r + 120)):
        if _to_text(text_rows[i][1] if len(text_rows[i]) > 1 else "").startswith("Adjustments"):
            adjustments_header = i
            break
    if adjustments_header >= 0:
        ar = adjustments_header + 2
        while ar < len(text_rows):
            name = _to_text(text_rows[ar][1] if len(text_rows[ar]) > 1 else "")
            if not name:
                break
            out.adjustments.append(
                MonthlyAdjustment(
                    Name=name,
                    Type=_to_text(text_rows[ar][2] if len(text_rows[ar]) > 2 else ""),
                    Value=_parse_decimal(text_rows[ar][3] if len(text_rows[ar]) > 3 else ""),
                )
            )
            ar += 1
        r = ar

    totals_header = -1
    for i in range(r, min(len(text_rows), r + 150)):
        if _to_text(text_rows[i][1] if len(text_rows[i]) > 1 else "") == "Category":
            totals_header = i
            break
    if totals_header >= 0:
        tr = totals_header + 1
        while tr < len(text_rows):
            cat = _to_text(text_rows[tr][1] if len(text_rows[tr]) > 1 else "")
            if not cat:
                break
            total = MonthlyEmployeeTotal(
                Category=cat,
                BasicHours=_parse_decimal(text_rows[tr][3] if len(text_rows[tr]) > 3 else ""),
                MonFriOvertime=_parse_decimal(text_rows[tr][4] if len(text_rows[tr]) > 4 else ""),
                SatSunOvertime=_parse_decimal(text_rows[tr][5] if len(text_rows[tr]) > 5 else ""),
                AnnualHoliday=_parse_decimal(text_rows[tr][6] if len(text_rows[tr]) > 6 else ""),
                TotalPaidHours=_parse_decimal(text_rows[tr][7] if len(text_rows[tr]) > 7 else ""),
            )
            out.employee_totals.append(total)
            tr += 1

    _enrich_week_summary(out)
    return out


def parse_monthly_inputs(
    weekly_files: list[Any],
    week_dates: list[tuple[str, str]] | None = None,
) -> list[MonthlyWeekSummary]:
    week_dates = week_dates or []
    summaries: list[MonthlyWeekSummary] = []
    for i, f in enumerate(weekly_files):
        start_date, end_date = week_dates[i] if i < len(week_dates) else ("", "")
        f.seek(0)
        if _workbook_has_all_data(f):
            f.seek(0)
            s = parse_weekly_gazebo_all_data(f, start_date=start_date, end_date=end_date)
        else:
            f.seek(0)
            s = parse_monthly_week_file(f)
            if start_date:
                s.start_date = start_date
            if end_date:
                s.end_date = end_date
        summaries.append(s)
    return summaries


def monthly_summaries_to_json(summaries: list[MonthlyWeekSummary]) -> list[dict[str, Any]]:
    return [asdict(s) for s in summaries]


def monthly_summaries_from_json(data: list[dict[str, Any]]) -> list[MonthlyWeekSummary]:
    out: list[MonthlyWeekSummary] = []
    for d in data:
        s = MonthlyWeekSummary(
            employees=[MonthlyEmployee(**e) for e in d.get("employees", [])],
            employee_totals=[MonthlyEmployeeTotal(**t) for t in d.get("employee_totals", [])],
            adjustments=[MonthlyAdjustment(**a) for a in d.get("adjustments", [])],
            start_date=str(d.get("start_date", "")),
            end_date=str(d.get("end_date", "")),
            non_agency_total=float(d.get("non_agency_total", 0.0)),
            emp_agency_bands=dict(d.get("emp_agency_bands") or {}),
            total_extra_hours=float(d.get("total_extra_hours", 0.0)),
            total_additional_holiday_pay=float(d.get("total_additional_holiday_pay", 0.0)),
        )
        s.grouped_totals = {}
        for k, v in (d.get("grouped_totals") or {}).items():
            if isinstance(v, dict):
                s.grouped_totals[str(k)] = MonthlyEmployeeTotal(**v)
        if not s.emp_agency_bands and s.employees:
            _enrich_week_summary(s)
        out.append(s)
    return out


def _set_column_widths(ws) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    for col in ("D", "E", "F", "G", "H", "I", "J", "K"):
        ws.column_dimensions[col].width = 12


def _apply_table_style(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    header_row: int | None = None,
) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = _MONTHLY_BORDER
            if header_row is not None and r == header_row:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif isinstance(cell.value, (int, float)) or (
                isinstance(cell.value, str) and str(cell.value).startswith("=")
            ):
                if c == _PAY_ID_COL and isinstance(cell.value, (int, float)):
                    cell.number_format = _INTEGER_FORMAT
                else:
                    cell.number_format = _NUM_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")


def _write_sheet_banner(ws, title: str, start_date: str, end_date: str) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_SHEET_LAST_COL)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = _SHEET_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    period_parts: list[str] = []
    if start_date:
        period_parts.append(f"D {start_date}")
    if end_date:
        period_parts.append(f"D {end_date}")
    period_text = f"Period: {' to '.join(period_parts)}" if period_parts else ""
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_SHEET_LAST_COL)
    period_cell = ws.cell(2, 1, period_text)
    period_cell.font = _SECTION_SUBTITLE_FONT
    period_cell.alignment = Alignment(horizontal="left", vertical="center")

    _set_column_widths(ws)
    return 4


def _write_section_title(ws, r: int, title: str, subtitle: str = "") -> int:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_SHEET_LAST_COL)
    ws.cell(r, 1, title).font = _SECTION_TITLE_FONT
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1
    if subtitle:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_SHEET_LAST_COL)
        ws.cell(r, 1, subtitle).font = _SECTION_SUBTITLE_FONT
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1
    return r


def _band_col_index(band_index: int) -> int:
    """0-based index into _BAND_KEYS -> Excel column (D=4)."""
    return 4 + band_index


def _employee_value_col_index(value_index: int) -> int:
    """0-based index into _EMPLOYEE_VALUE_KEYS -> Excel column (D=4)."""
    return 4 + value_index


_EXTRA_HOURS_COL = _employee_value_col_index(_EMPLOYEE_VALUE_KEYS.index("ExtraHours"))
_HOLIDAY_PAY_COL = _employee_value_col_index(_EMPLOYEE_VALUE_KEYS.index("AdditionalHolidayPay"))


def _xl_sumifs_category(row: int, data_start: int, data_end: int, band_col: int) -> str:
    col = get_column_letter(band_col)
    if data_end < data_start:
        return "=0"
    return (
        f"=SUMIFS(${col}${data_start}:${col}${data_end},"
        f"$B${data_start}:$B${data_end},$B{row})"
    )


def _xl_sumproduct_agency(data_start: int, data_end: int, band_col: int) -> str:
    col = get_column_letter(band_col)
    if data_end < data_start:
        return "=0"
    return (
        f'=SUMPRODUCT((LEFT($B${data_start}:$B${data_end},2)="A-")*'
        f"(${col}${data_start}:${col}${data_end}))"
    )


def _xl_emp_band(data_start: int, data_end: int, band_col: int, agency_row: int) -> str:
    col = get_column_letter(band_col)
    if data_end < data_start:
        return "=0"
    return f"=SUM(${col}${data_start}:${col}${data_end})-{col}{agency_row}"


def _xl_total_band(emp_row: int, agency_row: int, band_col: int) -> str:
    col = get_column_letter(band_col)
    return f"={col}{emp_row}+{col}{agency_row}"


def _xl_cross_week_sumif(name_cell: str, week_sheet_names: list[str], band_col: int) -> str:
    col = get_column_letter(band_col)
    parts = [
        f"SUMIF('{sheet}'!$A:$A,{name_cell},'{sheet}'!${col}:${col})"
        for sheet in week_sheet_names
    ]
    return "=" + "+".join(parts)


def _xl_sum_week_emp_cells(week_layouts: list[WeekSheetLayout], row_key: str, band_col: int) -> str:
    col = get_column_letter(band_col)
    parts = [
        f"'{layout.sheet_name}'!{col}{layout.emp_agency.rows[row_key]}"
        for layout in week_layouts
    ]
    return "=" + "+".join(parts)


def _write_employee_header(ws, r: int) -> None:
    labels = (
        "Name",
        "Category",
        "Pay ID (Sage)",
        *_BAND_LABELS,
        *_HOLIDAY_PAY_LABELS,
    )
    for i, label in enumerate(labels):
        ws.cell(r, 1 + i, label)


def _write_holiday_pay_header(ws, r: int) -> None:
    ws.cell(r, 2, "Category")
    ws.cell(r, _EXTRA_HOURS_COL, "Extra hours")
    ws.cell(r, _HOLIDAY_PAY_COL, "Additional holiday pay")


def _write_total_header(ws, r: int) -> None:
    ws.cell(r, 2, "Category")
    for j, label in enumerate(_BAND_LABELS):
        ws.cell(r, 4 + j, label)


def _write_adjustment_header(ws, r: int) -> None:
    ws.cell(r, 2, "Name")
    ws.cell(r, 3, "Type")
    ws.cell(r, 4, "Value")


def _write_employee_row(ws, r: int, e: MonthlyEmployee) -> None:
    ws.cell(r, 1, e.Name)
    ws.cell(r, 2, e.Category)
    ws.cell(r, 3, int(e.SageNo))
    values = (
        e.BasicHours,
        e.MonFriOvertime,
        e.SatSunOvertime,
        e.AnnualHoliday,
        e.TotalPaidHours,
        e.ContractedHours,
        e.ExtraHours,
        e.AdditionalHolidayPay,
    )
    for i, value in enumerate(values):
        ws.cell(r, _employee_value_col_index(i), value)


def _write_total_row(ws, r: int, category: str, t: MonthlyEmployeeTotal) -> None:
    ws.cell(r, 2, category)
    ws.cell(r, 4, t.BasicHours)
    ws.cell(r, 5, t.MonFriOvertime)
    ws.cell(r, 6, t.SatSunOvertime)
    ws.cell(r, 7, t.AnnualHoliday)
    ws.cell(r, 8, t.TotalPaidHours)


def _write_emp_agency_block_formulas(
    ws,
    r: int,
    emp_layout: EmployeeTableLayout,
    *,
    row_label: str | None = None,
) -> tuple[int, dict[str, int]]:
    """Write EMP/AGENCY/TOTAL with SUMPRODUCT/SUM formulas from employee table."""
    block_start = r
    ds, de = emp_layout.data_start, emp_layout.data_end
    emp_row = r
    agency_row = r + 1
    total_row = r + 2

    if row_label:
        ws.cell(r, 2, row_label)

    ws.cell(emp_row, 3, "EMP")
    ws.cell(agency_row, 3, "AGENCY")
    ws.cell(total_row, 3, "TOTAL")

    for j in range(5):
        col = _band_col_index(j)
        ws.cell(agency_row, col, _xl_sumproduct_agency(ds, de, col))
        ws.cell(emp_row, col, _xl_emp_band(ds, de, col, agency_row))
        ws.cell(total_row, col, _xl_total_band(emp_row, agency_row, col))

    row_map = {"EMP": emp_row, "AGENCY": agency_row, "TOTAL": total_row}
    _apply_table_style(ws, block_start, total_row, 2, _SHEET_LAST_COL)
    return total_row + 1, row_map


def _write_holiday_pay_block_formulas(
    ws,
    r: int,
    emp_layout: EmployeeTableLayout,
    *,
    row_label: str | None = None,
) -> tuple[int, dict[str, int]]:
    """EMP/AGENCY/TOTAL for extra hours and additional holiday pay (cols J–K)."""
    block_start = r
    ds, de = emp_layout.data_start, emp_layout.data_end
    emp_row = r
    agency_row = r + 1
    total_row = r + 2

    if row_label:
        ws.cell(r, 2, row_label)

    ws.cell(emp_row, 3, "EMP")
    ws.cell(agency_row, 3, "AGENCY")
    ws.cell(total_row, 3, "TOTAL")

    for col in (_EXTRA_HOURS_COL, _HOLIDAY_PAY_COL):
        ws.cell(agency_row, col, _xl_sumproduct_agency(ds, de, col))
        ws.cell(emp_row, col, _xl_emp_band(ds, de, col, agency_row))
        ws.cell(total_row, col, _xl_total_band(emp_row, agency_row, col))

    row_map = {"EMP": emp_row, "AGENCY": agency_row, "TOTAL": total_row}
    _apply_table_style(ws, block_start, total_row, 2, _SHEET_LAST_COL)
    return total_row + 1, row_map


def _write_holiday_pay_section_formulas(
    ws,
    r: int,
    emp_layout: EmployeeTableLayout,
    section_title: str,
    subtitle: str,
    *,
    row_label: str | None = None,
) -> HolidayPayLayout:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_holiday_pay_header(ws, r)
    r += 1
    r, row_map = _write_holiday_pay_block_formulas(ws, r, emp_layout, row_label=row_label)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return HolidayPayLayout(header_row=header_row, rows=row_map, next_row=r + 1)


def _write_holiday_pay_block_week_refs(
    ws,
    r: int,
    week_layout: WeekSheetLayout,
    *,
    row_label: str | None = None,
) -> int:
    block_start = r
    for i, key in enumerate(_EMP_AGENCY_ROWS):
        if row_label and i == 0:
            ws.cell(r, 2, row_label)
        ws.cell(r, 3, key)
        src_row = week_layout.holiday_pay.rows[key]
        for col in (_EXTRA_HOURS_COL, _HOLIDAY_PAY_COL):
            ws.cell(r, col, f"='{week_layout.sheet_name}'!{get_column_letter(col)}{src_row}")
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    return r


def _write_holiday_pay_section_monthly_formulas(
    ws,
    r: int,
    week_layouts: list[WeekSheetLayout],
    section_title: str,
    subtitle: str,
) -> tuple[int, HolidayPayLayout]:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_holiday_pay_header(ws, r)
    r += 1
    block_start = r
    ws.cell(r, 2, "MONTHLY")
    row_map: dict[str, int] = {}
    for key in _EMP_AGENCY_ROWS:
        row_map[key] = r
        ws.cell(r, 3, key)
        for col in (_EXTRA_HOURS_COL, _HOLIDAY_PAY_COL):
            ws.cell(
                r,
                col,
                "="
                + "+".join(
                    f"'{layout.sheet_name}'!{get_column_letter(col)}{layout.holiday_pay.rows[key]}"
                    for layout in week_layouts
                ),
            )
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return r + 1, HolidayPayLayout(header_row=header_row, rows=row_map, next_row=r + 1)


def _write_emp_agency_block(
    ws,
    r: int,
    bands: dict[str, dict[str, float]],
    *,
    row_label: str | None = None,
) -> int:
    """EMP/AGENCY/TOTAL in col C with bands in D–H; optional label in col B on first row."""
    block_start = r
    for i, key in enumerate(_EMP_AGENCY_ROWS):
        if row_label and i == 0:
            ws.cell(r, 2, row_label)
        ws.cell(r, 3, key)
        row_bands = bands.get(key) or _empty_bands()
        for j, col in enumerate(_BAND_KEYS):
            ws.cell(r, 4 + j, float(row_bands.get(col, 0.0) or 0.0))
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    return r


def _write_emp_agency_section_formulas(
    ws,
    r: int,
    emp_layout: EmployeeTableLayout,
    section_title: str,
    subtitle: str,
    *,
    row_label: str | None = None,
) -> EmpAgencyLayout:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    r, row_map = _write_emp_agency_block_formulas(ws, r, emp_layout, row_label=row_label)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return EmpAgencyLayout(header_row=header_row, rows=row_map, next_row=r + 1)


def _write_emp_agency_section_monthly_formulas(
    ws,
    r: int,
    week_layouts: list[WeekSheetLayout],
    section_title: str,
    subtitle: str,
) -> tuple[int, EmpAgencyLayout]:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    block_start = r
    ws.cell(r, 2, "MONTHLY")
    row_map: dict[str, int] = {}
    for i, key in enumerate(_EMP_AGENCY_ROWS):
        row_map[key] = r
        ws.cell(r, 3, key)
        for j in range(5):
            ws.cell(r, _band_col_index(j), _xl_sum_week_emp_cells(week_layouts, key, _band_col_index(j)))
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return r + 1, EmpAgencyLayout(header_row=header_row, rows=row_map, next_row=r + 1)


def _write_emp_agency_block_week_refs(
    ws,
    r: int,
    week_layout: WeekSheetLayout,
    *,
    row_label: str | None = None,
) -> int:
    """Per-week block on Summary referencing Week sheet EMP/AGENCY/TOTAL cells."""
    block_start = r
    for i, key in enumerate(_EMP_AGENCY_ROWS):
        if row_label and i == 0:
            ws.cell(r, 2, row_label)
        ws.cell(r, 3, key)
        src_row = week_layout.emp_agency.rows[key]
        for j in range(5):
            col = _band_col_index(j)
            ws.cell(r, col, f"='{week_layout.sheet_name}'!{get_column_letter(col)}{src_row}")
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    return r


def _write_emp_agency_section_diff_formulas(
    ws,
    r: int,
    summary_emp_layout: EmployeeTableLayout,
    monthly_layout: EmpAgencyLayout,
    section_title: str,
    subtitle: str,
) -> int:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    ds, de = summary_emp_layout.data_start, summary_emp_layout.data_end

    emp_merged_row = r
    ws.cell(emp_merged_row, 3, "EMP")
    agency_merged_row = r + 1
    ws.cell(agency_merged_row, 3, "AGENCY")
    total_merged_row = r + 2
    ws.cell(total_merged_row, 3, "TOTAL")
    ws.cell(emp_merged_row, 2, "Merged (employees)")

    for j in range(5):
        col = _band_col_index(j)
        ws.cell(agency_merged_row, col, _xl_sumproduct_agency(ds, de, col))
        ws.cell(emp_merged_row, col, _xl_emp_band(ds, de, col, agency_merged_row))
        ws.cell(total_merged_row, col, _xl_total_band(emp_merged_row, agency_merged_row, col))

    merged_rows = {"EMP": emp_merged_row, "AGENCY": agency_merged_row, "TOTAL": total_merged_row}
    diff_row_start = r + 3
    ws.cell(diff_row_start, 2, "Diff")

    for i, key in enumerate(_EMP_AGENCY_ROWS):
        row = diff_row_start + i
        ws.cell(row, 3, key)
        for j in range(5):
            col = _band_col_index(j)
            cl = get_column_letter(col)
            ws.cell(row, col, f"={cl}{merged_rows[key]}-{cl}{monthly_layout.rows[key]}")
        r = row + 1

    _apply_table_style(ws, header_row, r - 1, 2, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_employee_table_cross_week(
    ws,
    r: int,
    employees: list[MonthlyEmployee],
    section_title: str,
    subtitle: str,
    week_sheet_names: list[str],
) -> EmployeeTableLayout:
    r = _write_section_title(
        ws,
        r,
        section_title,
        subtitle + " Totals use Excel formulas — click any hour cell to see Week sheet references.",
    )
    header_row = r
    _write_employee_header(ws, r)
    r += 1
    data_start = r
    for e in employees:
        ws.cell(r, 1, e.Name)
        ws.cell(r, 2, e.Category)
        ws.cell(r, 3, int(e.SageNo))
        name_ref = f"$A{r}"
        for j in range(len(_EMPLOYEE_VALUE_KEYS)):
            ws.cell(r, _employee_value_col_index(j), _xl_cross_week_sumif(name_ref, week_sheet_names, _employee_value_col_index(j)))
        r += 1
    data_end = r - 1 if employees else data_start - 1
    if employees:
        _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 1, _SHEET_LAST_COL, header_row=header_row)
    return EmployeeTableLayout(header_row=header_row, data_start=data_start, data_end=data_end, next_row=r + 1)


def _write_employee_table_source(
    ws,
    r: int,
    employees: list[MonthlyEmployee],
    section_title: str,
    subtitle: str,
) -> EmployeeTableLayout:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_employee_header(ws, r)
    r += 1
    data_start = r
    for e in employees:
        _write_employee_row(ws, r, e)
        r += 1
    data_end = r - 1 if employees else data_start - 1
    if employees:
        _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 1, _SHEET_LAST_COL, header_row=header_row)
    return EmployeeTableLayout(header_row=header_row, data_start=data_start, data_end=data_end, next_row=r + 1)


def _write_totals_table_formulas(
    ws,
    r: int,
    categories: list[str],
    emp_layout: EmployeeTableLayout,
    section_title: str,
    subtitle: str,
) -> int:
    if not categories:
        return r
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    data_start = r
    ds, de = emp_layout.data_start, emp_layout.data_end
    for category in categories:
        ws.cell(r, 2, category)
        for j in range(5):
            ws.cell(r, _band_col_index(j), _xl_sumifs_category(r, ds, de, _band_col_index(j)))
        r += 1
    _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_totals_table(
    ws,
    r: int,
    rows: list[tuple[str, MonthlyEmployeeTotal]],
    section_title: str,
    subtitle: str,
) -> int:
    if not rows:
        return r
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    for category, t in rows:
        _write_total_row(ws, r, category, t)
        r += 1
    _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_adjustments_table(ws, r: int, adjustments: list[MonthlyAdjustment]) -> int:
    r = _write_section_title(
        ws,
        r,
        f"Adjustments ({len(adjustments)})",
        "Manual hour corrections recorded in the legacy weekly file.",
    )
    header_row = r
    _write_adjustment_header(ws, r)
    r += 1
    for a in adjustments:
        ws.cell(r, 2, a.Name)
        ws.cell(r, 3, a.Type)
        ws.cell(r, 4, a.Value)
        r += 1
    if adjustments:
        _apply_table_style(ws, header_row, r - 1, 2, 4, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 2, 4, header_row=header_row)
    return r + 1


def _write_emp_agency_section(
    ws,
    r: int,
    bands: dict[str, dict[str, float]],
    section_title: str,
    subtitle: str,
    *,
    row_label: str | None = None,
) -> int:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    r = _write_emp_agency_block(ws, r, bands, row_label=row_label)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_summary_category_formulas(
    ws,
    r: int,
    categories: list[str],
    emp_layout: EmployeeTableLayout,
) -> int:
    r = _write_section_title(
        ws,
        r,
        "Category totals (month)",
        "Sum of hours by work category — calculated with SUMIFS from the employee table above.",
    )
    header_row = r
    _write_total_header(ws, r)
    r += 1
    cat_data_start = r
    ds, de = emp_layout.data_start, emp_layout.data_end
    for category in categories:
        ws.cell(r, 2, category)
        for j in range(5):
            ws.cell(r, _band_col_index(j), _xl_sumifs_category(r, ds, de, _band_col_index(j)))
        r += 1
    if r > cat_data_start:
        _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 1, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def build_monthly_excel_bytes(
    week_summaries: list[MonthlyWeekSummary],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    grouped_from_weeks: list[dict[str, MonthlyEmployeeTotal]] = []
    week_layouts: list[WeekSheetLayout] = []

    for i, s in enumerate(week_summaries, start=1):
        if not s.emp_agency_bands:
            _enrich_week_summary(s)
        sheet_name = f"Week{i}"
        ws = wb.create_sheet(sheet_name)
        r = _write_sheet_banner(ws, f"Gazebo HR — Week {i}", s.start_date, s.end_date)
        emp_layout = _write_employee_table_source(
            ws,
            r,
            s.employees,
            f"Week {i} — Employee paid hours",
            "One row per employee for this pay week (from weekly export).",
        )
        r = emp_layout.next_row
        r = _write_adjustments_table(ws, r, s.adjustments)
        categories = [t.Category for t in s.employee_totals]
        r = _write_totals_table_formulas(
            ws,
            r,
            categories,
            emp_layout,
            "Category totals",
            "Sum of hours by work category for this week (SUMIFS from employee rows above).",
        )
        grouped_rows = [(k, s.grouped_totals[k]) for k in sorted(s.grouped_totals.keys())]
        r = _write_totals_table(
            ws,
            r,
            grouped_rows,
            "Grouped totals (by category prefix)",
            "Categories rolled up to 4-character groups.",
        )
        overall_rows = _build_overall_totals_from_employee_totals(s.employee_totals)
        r = _write_totals_table(
            ws,
            r,
            overall_rows,
            "Category breakdown (overall)",
            "Hours by department — Gazebo and agency combined (PROD, PACK, WRHS, CLNR, TECH, OFFICE).",
        )
        emp_agency = _write_emp_agency_section_formulas(
            ws,
            r,
            emp_layout,
            "Gazebo vs agency summary",
            "EMP = Gazebo staff; AGENCY = A- categories; TOTAL = both (formulas from employee table).",
        )
        holiday_pay = _write_holiday_pay_section_formulas(
            ws,
            emp_agency.next_row,
            emp_layout,
            "Additional holiday pay",
            "Extra hours = actual − contracted; additional holiday pay = 0.1207 × extra hours (weekly report rules).",
        )
        week_layouts.append(
            WeekSheetLayout(
                sheet_name=sheet_name,
                employee=emp_layout,
                emp_agency=emp_agency,
                holiday_pay=holiday_pay,
            )
        )
        grouped_from_weeks.append(s.grouped_totals)

    merged_employees: dict[str, MonthlyEmployee] = {}
    merged_totals: dict[str, MonthlyEmployeeTotal] = {}
    for s in week_summaries:
        for e in s.employees:
            cur = merged_employees.get(e.Name)
            if cur is None:
                merged_employees[e.Name] = MonthlyEmployee(**e.__dict__)
                continue
            cur.BasicHours += e.BasicHours
            cur.MonFriOvertime += e.MonFriOvertime
            cur.SatSunOvertime += e.SatSunOvertime
            cur.AnnualHoliday += e.AnnualHoliday
            cur.TotalPaidHours += e.TotalPaidHours
            cur.ContractedHours += e.ContractedHours
            cur.ExtraHours += e.ExtraHours
            cur.AdditionalHolidayPay += e.AdditionalHolidayPay
            cur.SageNo = e.SageNo
            cur.Category = e.Category

        for t in s.employee_totals:
            cur_t = merged_totals.get(t.Category)
            if cur_t is None:
                merged_totals[t.Category] = MonthlyEmployeeTotal(**t.__dict__)
                continue
            cur_t.BasicHours += t.BasicHours
            cur_t.MonFriOvertime += t.MonFriOvertime
            cur_t.SatSunOvertime += t.SatSunOvertime
            cur_t.AnnualHoliday += t.AnnualHoliday
            cur_t.TotalPaidHours += t.TotalPaidHours

    week_sheet_names = [wl.sheet_name for wl in week_layouts]
    summary_start = week_summaries[0].start_date if week_summaries else ""
    summary_end = week_summaries[-1].end_date if week_summaries else ""
    ws = wb.create_sheet("Summary")
    r = _write_sheet_banner(ws, "Gazebo HR — Monthly Summary", summary_start, summary_end)
    summary_emp_layout = _write_employee_table_cross_week(
        ws,
        r,
        list(merged_employees.values()),
        "Monthly summary — all employees (weeks combined)",
        "Total paid hours per employee across all uploaded weeks.",
        week_sheet_names,
    )
    r = summary_emp_layout.next_row
    r = _write_summary_category_formulas(
        ws,
        r,
        list(merged_totals.keys()),
        summary_emp_layout,
    )

    agg_grouped: dict[str, MonthlyEmployeeTotal] = {}
    for m in grouped_from_weeks:
        for k, t in m.items():
            if k not in agg_grouped:
                agg_grouped[k] = MonthlyEmployeeTotal(k, 0.0, 0.0, 0.0, 0.0, 0.0)
            a = agg_grouped[k]
            a.BasicHours += t.BasicHours
            a.MonFriOvertime += t.MonFriOvertime
            a.SatSunOvertime += t.SatSunOvertime
            a.AnnualHoliday += t.AnnualHoliday
            a.TotalPaidHours += t.TotalPaidHours
    grouped_rows = [(k, agg_grouped[k]) for k in sorted(agg_grouped.keys())]
    r = _write_totals_table(
        ws,
        r,
        grouped_rows,
        "Grouped totals (month)",
        "Categories rolled up to 4-character groups across all weeks.",
    )
    overall_rows = _build_overall_totals_from_employee_totals(list(merged_totals.values()))
    r = _write_totals_table(
        ws,
        r,
        overall_rows,
        "Category breakdown (overall)",
        "Month total by department — Gazebo and agency combined (PROD, PACK, WRHS, CLNR, TECH, OFFICE).",
    )

    if week_layouts:
        r, monthly_layout = _write_emp_agency_section_monthly_formulas(
            ws,
            r,
            week_layouts,
            "Month total — Gazebo vs agency",
            "Sum of weekly EMP/AGENCY/TOTAL rows across all uploaded weeks.",
        )

        r, _monthly_holiday_layout = _write_holiday_pay_section_monthly_formulas(
            ws,
            r,
            week_layouts,
            "Month total — additional holiday pay",
            "Sum of weekly extra hours and additional holiday pay (EMP / AGENCY / TOTAL).",
        )

        holiday_pay = _write_holiday_pay_section_formulas(
            ws,
            r,
            summary_emp_layout,
            "Merged additional holiday pay",
            "Recalculated from merged employee rows using the same weekly rules.",
        )

        r = _write_section_title(
            ws,
            r,
            "Per-week Gazebo vs agency",
            "Each row references the matching EMP/AGENCY/TOTAL cells on the Week sheet.",
        )
        for wi, layout in enumerate(week_layouts):
            label = "Weekly" if wi == 0 else None
            header_row = r
            _write_total_header(ws, r)
            r += 1
            r = _write_emp_agency_block_week_refs(ws, r, layout, row_label=label)
            _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
            r += 1

        r = _write_section_title(
            ws,
            r,
            "Per-week additional holiday pay",
            "Each row references the matching EMP/AGENCY/TOTAL cells on the Week sheet.",
        )
        for wi, layout in enumerate(week_layouts):
            label = "Weekly" if wi == 0 else None
            header_row = r
            _write_holiday_pay_header(ws, r)
            r += 1
            r = _write_holiday_pay_block_week_refs(ws, r, layout, row_label=label)
            _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
            r += 1

        r = _write_emp_agency_section_diff_formulas(
            ws,
            r,
            summary_emp_layout,
            monthly_layout,
            "Reconciliation difference",
            "Merged (employees) minus MONTHLY row — EMP, AGENCY and TOTAL should all be zero.",
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
